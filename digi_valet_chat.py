"""
Digi Valet — Local AI Personal Assistant
Built with PySide6 + Ollama
Run: python digi_valet_chat.py

Enhanced Edition v2 — Features:
  • Persistent multi-session chat history
  • Enter to send / Shift+Enter for new line
  • Tone of voice selector (Formal / Balanced / Casual)
  • Language selector (multilingual system prompts)
  • Quick-command shortcuts (/help, /tasks, /plan, /wellness, /meal)
  • Sidebar tasks tracker (personal task management)
  • Privacy mode (history not saved)
  • Light / Dark theme toggle
  • ★ Universal file attachment (any file type, no size/count limit)
  • ★ Auto file-context training (files injected into session context)
  • ★ Claude-style thinking steps (collapsible, expandable process cards)
  • ★ Code Preview Panel (Preview + Code tabs, like Claude artifacts)
"""

import sys
import json
import re
import urllib.request
import urllib.error
import threading
import os
import mimetypes
import base64
import zipfile
import csv as csv_module
import struct
from datetime import datetime
from pathlib import Path


# ── OrgTree (fully embedded — no external file needed) ────────────────────────
class OrgTree:
    """
    Embedded org-chart lookup backed by OrgTree.csv.
    Columns expected: Employee Number, Display Name, Job Title,
                      Department, Location, Reportees Count
    """
    loaded = False
    source_path = None

    def __init__(self):
        self.employees: dict = {}   # emp_number.lower() → record dict
        self._name_index: dict = {} # display_name.lower() → emp_number

    def load_from_csv(self, path: str) -> str:
        import csv as _csv
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = _csv.DictReader(f)
                raw_headers = reader.fieldnames or []
                headers = [h.strip() for h in raw_headers]
                # Normalise common column name variants
                _col = {}
                for h in headers:
                    hl = h.lower().replace(' ', '_').replace('-', '_')
                    for canon, variants in {
                        'emp_number': ['employee_number', 'emp_no', 'emp_id', 'id'],
                        'name':       ['display_name', 'full_name', 'employee_name', 'name'],
                        'title':      ['job_title', 'designation', 'role', 'position'],
                        'dept':       ['department', 'dept', 'division', 'team'],
                        'location':   ['location', 'city', 'office', 'site'],
                        'reports':    ['reportees_count', 'reportees', 'direct_reports', 'reports'],
                    }.items():
                        if hl in variants or canon in hl:
                            _col[canon] = h
                            break

                self.employees.clear()
                self._name_index.clear()
                for row in reader:
                    row = {k.strip(): (v or '').strip() for k, v in row.items()}
                    emp_no  = row.get(_col.get('emp_number',''), '').strip()
                    name    = row.get(_col.get('name',''), '').strip()
                    title   = row.get(_col.get('title',''), '').strip()
                    dept    = row.get(_col.get('dept',''), '').strip()
                    loc     = row.get(_col.get('location',''), '').strip()
                    reports = row.get(_col.get('reports',''), '0').strip()
                    if not (emp_no or name):
                        continue
                    try:
                        rc = int(str(reports).split('.')[0])
                    except (ValueError, AttributeError):
                        rc = 0
                    key = (emp_no or name).lower()
                    record = {
                        'emp_number': emp_no, 'name': name, 'title': title,
                        'dept': dept, 'location': loc, 'reports_count': rc
                    }
                    self.employees[key] = record
                    if name:
                        self._name_index[name.lower()] = key

            self.loaded = True
            self.source_path = path
            return f"✓ OrgTree loaded: {len(self.employees)} employees from {path}"
        except Exception as e:
            self.loaded = False
            return f"⚠️ OrgTree load error: {e}"

    def get_employee(self, query: str):
        q = query.strip().lower()
        if q in self.employees:
            return self.employees[q]
        if q in self._name_index:
            return self.employees[self._name_index[q]]
        return None

    def search(self, query: str, limit: int = 30) -> list:
        q = query.strip().lower()
        results = []
        for rec in self.employees.values():
            haystack = ' '.join([
                rec.get('emp_number',''), rec.get('name',''),
                rec.get('title',''), rec.get('dept',''), rec.get('location','')
            ]).lower()
            if q in haystack:
                results.append(rec)
                if len(results) >= limit:
                    break
        return results

    def list_by_department(self, dept: str, limit: int = 50) -> list:
        d = dept.strip().lower()
        return [r for r in self.employees.values() if d in r.get('dept','').lower()][:limit]

    def list_by_location(self, loc: str, limit: int = 50) -> list:
        l = loc.strip().lower()
        return [r for r in self.employees.values() if l in r.get('location','').lower()][:limit]

    def stats(self) -> dict:
        from collections import Counter
        depts = Counter(r.get('dept','—') for r in self.employees.values())
        locs  = Counter(r.get('location','—') for r in self.employees.values())
        return {
            'total': len(self.employees),
            'departments': dict(depts.most_common(30)),
            'locations':   dict(locs.most_common(20)),
        }

    def to_context_snippet(self, max_rows: int = 200) -> str:
        if not self.loaded:
            return ""
        lines = ["Emp# | Name | Title | Dept | Location | Reports"]
        for rec in list(self.employees.values())[:max_rows]:
            lines.append(
                f"{rec.get('emp_number','')} | {rec.get('name','')} | "
                f"{rec.get('title','')} | {rec.get('dept','')} | "
                f"{rec.get('location','')} | {rec.get('reports_count',0)}"
            )
        return '\n'.join(lines)

    @staticmethod
    def format_employee(info: dict) -> str:
        if not info:
            return "Employee not found."
        return (
            f"**{info.get('name','—')}**  (`{info.get('emp_number','—')}`)\n\n"
            f"- **Title:** {info.get('title','—')}\n"
            f"- **Department:** {info.get('dept','—')}\n"
            f"- **Location:** {info.get('location','—')}\n"
            f"- **Direct Reports:** {info.get('reports_count', 0)}"
        )

    @staticmethod
    def format_results_table(results: list) -> str:
        if not results:
            return "No results."
        lines = ["| Emp# | Name | Title | Dept | Location | Reports |",
                 "|---|---|---|---|---|---|"]
        for r in results:
            lines.append(
                f"| {r.get('emp_number','')} | {r.get('name','')} | "
                f"{r.get('title','')} | {r.get('dept','')} | "
                f"{r.get('location','')} | {r.get('reports_count',0)} |"
            )
        return '\n'.join(lines)

_ORG_TREE_AVAILABLE = True   # always available — it's embedded above

# ── KnowledgeBase (embedded) ──────────────────────────────────────────────────
class KnowledgeBase:
    """Persistent per-user knowledge store backed by a JSON file."""
    def __init__(self, kb_path=None):
        from pathlib import Path as _Path
        self._path = _Path(kb_path) if kb_path else _Path.home() / ".digi_valet_kb.json"
        self._data: dict = self._load()

    def _load(self) -> dict:
        try:
            if self._path.exists():
                import json as _j
                return _j.loads(self._path.read_text(encoding='utf-8'))
        except Exception:
            pass
        return {"files": {}, "notes": []}

    def _save(self):
        import json as _j
        try:
            self._path.write_text(_j.dumps(self._data, indent=2, ensure_ascii=False), encoding='utf-8')
        except Exception as e:
            print(f"KB save error: {e}")

    def is_empty(self) -> bool:
        return not self._data.get("files") and not self._data.get("notes")

    def list_files(self) -> list:
        return [{"filename": k, **v} for k, v in self._data.get("files", {}).items()]

    def learn_file(self, filepath: str) -> str:
        from pathlib import Path as _Path
        p = _Path(filepath)
        _, content = _read_file_content(filepath)
        self._data.setdefault("files", {})[p.name] = {
            "path": str(p), "content": content,
            "learned_at": datetime.now().isoformat()
        }
        self._save()
        return f"✓ Learned: {p.name}"

    def forget_file(self, filename: str) -> str:
        if filename in self._data.get("files", {}):
            del self._data["files"][filename]
            self._save()
            return f"✓ Forgotten: {filename}"
        return f"File '{filename}' not found in knowledge base."

    def system_prompt_block(self) -> str:
        lines = ["# Persistent Knowledge Base\n"]
        for fname, info in self._data.get("files", {}).items():
            lines.append(f"## {fname}\n{info.get('content','')[:4000]}\n")
        for note in self._data.get("notes", [])[:20]:
            lines.append(f"- {note}")
        return '\n'.join(lines)

    def format_kb_status(self) -> str:
        """Return a human-readable summary of the knowledge base contents."""
        files = self._data.get("files", {})
        notes = self._data.get("notes", [])
        if not files and not notes:
            return (
                "📚 **Knowledge Base** — empty\n\n"
                "Attach a file and type `/kb learn` to store it permanently, "
                "or say **'learn this file'** after attaching."
            )
        lines = [f"📚 **Knowledge Base** — {len(files)} file(s) stored\n"]
        for fname, info in files.items():
            size = len(info.get("content", ""))
            learned = info.get("learned_at", "")[:10]
            lines.append(f"• **{fname}** — {size:,} chars — learned {learned}")
        if notes:
            lines.append(f"\n{len(notes)} note(s) stored.")
        lines.append(
            "\n**Commands:** `/kb learn` · `/kb forget <filename>` · `/kb clear` · `/kb help`"
        )
        return "\n".join(lines)

# ── Excel Intelligence (embedded) ─────────────────────────────────────────────
try:
    from excel_intelligence import ExcelIntelligence as _ExcelIntelligence
    _excel_intel = _ExcelIntelligence()
    print("[Digi Valet] ✓ Excel Intelligence loaded")
except ImportError:
    _excel_intel = None

# ── Programming Knowledge Base (embedded directly) ────────────────────────────
PROGRAMMING_KNOWLEDGE = """
══════════════════════════════════════════════════════════════════════
PROGRAMMING & MACHINE LEARNING KNOWLEDGE BASE
You are an expert software engineer and ML scientist. Use this
knowledge to answer ANY programming or ML question accurately.
══════════════════════════════════════════════════════════════════════

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 1: PYTHON
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Python is a high-level, interpreted, dynamically typed language.
BASICS: variables (x=10), lists ([1,2,3]), dicts ({'a':1}), sets ({1,2,3}), tuples ((1,2,3)).
CONTROL FLOW: if/elif/else, for/while loops, list comprehensions [x**2 for x in range(10)].
FUNCTIONS: def greet(name, greeting="Hi"): return f"{greeting} {name}"
  lambda x: x*2   |   *args/**kwargs   |   decorators (@functools.wraps)
OOP: class Animal: def __init__(self,name): self.name=name  |  inheritance, super(), dunder methods
EXCEPTIONS: try/except/finally/raise  |  custom exceptions (class MyError(Exception))
FILE I/O: open('f','r') as f: f.read()  |  pathlib.Path
MODULES: import os,sys,json,re,math,datetime,collections,itertools,functools
ASYNC: async def fetch(): await asyncio.sleep(1)
POPULAR LIBS: numpy (arrays), pandas (DataFrames), matplotlib (plots), requests (HTTP),
  scikit-learn (ML), Flask/FastAPI (web), SQLAlchemy (ORM), pytest (testing)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 2: JAVA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Statically typed, JVM-based. public class Main { public static void main(String[] args){} }
Types: int, double, boolean, String, int[], ArrayList<>, HashMap<>
OOP: extends (inheritance), implements (interface), abstract classes, @Override
Exceptions: try/catch/finally, throws, checked vs unchecked
Collections: List, Set, Map, Queue, Iterator
Streams: list.stream().filter(x->x>0).map(x->x*2).collect(Collectors.toList())
Threads: Runnable, Thread, synchronized, ExecutorService
Java 8+: lambdas (x -> x*2), Optional, LocalDate, var (Java 10+), records (Java 16+)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 3: C / C++
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
C: #include <stdio.h>  int main(){ printf("hi"); return 0; }
  pointers (*ptr, &var), malloc/free, structs, arrays, string.h functions
C++: classes with constructors/destructors, new/delete, RAII
  STL: vector, map, set, queue, stack, algorithm (sort, find, etc.)
  Templates: template<typename T> T max(T a,T b){return a>b?a:b;}
  Smart pointers: unique_ptr, shared_ptr, weak_ptr
  Lambda: auto f = [](int x){ return x*2; };
  C++11/14/17/20: range-for, auto, constexpr, structured bindings, concepts

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 4: C# / .NET
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
using System; namespace App { class Program { static void Main(){} } }
Types: int,double,bool,string,var  |  nullable: int?  |  dynamic
OOP: public/private/protected, sealed, abstract, interface, partial
LINQ: list.Where(x=>x>0).Select(x=>x*2).ToList()
Async: async Task<int> GetData() { await Task.Delay(1000); return 42; }
Collections: List<T>, Dictionary<K,V>, HashSet<T>, IEnumerable<T>
.NET: Console, File/Directory, HttpClient, Entity Framework, ASP.NET Core

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 5: JAVASCRIPT / TYPESCRIPT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JS: var/let/const, arrow functions (x=>x*2), template literals (`${x}`)
  Array methods: map, filter, reduce, find, some, every, flat, flatMap
  Objects: destructuring ({a,b}=obj), spread ({...obj}), optional chaining (obj?.a)
  Async: Promise, async/await, fetch API, try/catch
  Classes: class Animal { constructor(name){this.name=name} speak(){} }
  Modules: import {fn} from './mod.js'  |  export default fn
DOM: document.querySelector, addEventListener, innerHTML, classList
Node.js: require/module.exports, fs, path, http, Express.js
TypeScript: type annotations (x:number), interfaces, generics <T>, enums, utility types

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 6: SQL & DATABASES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SELECT col FROM tbl WHERE cond GROUP BY col HAVING agg ORDER BY col LIMIT n
JOINs: INNER, LEFT, RIGHT, FULL OUTER, CROSS, SELF
Aggregates: COUNT(*), SUM(x), AVG(x), MAX/MIN(x), GROUP_CONCAT
Subqueries: SELECT * FROM t WHERE x IN (SELECT x FROM t2)
Window: ROW_NUMBER() OVER (PARTITION BY dept ORDER BY salary DESC)
Indexes: CREATE INDEX idx ON tbl(col)  — speeds up reads, slows writes
Transactions: BEGIN; UPDATE ...; COMMIT; / ROLLBACK;
NoSQL: MongoDB (documents/BSON), Redis (key-value), Cassandra (wide-column)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 7: MACHINE LEARNING
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SUPERVISED: regression (predict continuous), classification (predict category)
  Algorithms: Linear/Logistic Regression, Decision Tree, Random Forest, SVM,
              KNN, Naive Bayes, XGBoost/LightGBM, Neural Networks
UNSUPERVISED: clustering (K-Means, DBSCAN, hierarchical), dimensionality reduction (PCA, t-SNE, UMAP)
EVALUATION:
  Classification: accuracy, precision, recall, F1, ROC-AUC, confusion matrix
  Regression: MAE, MSE, RMSE, R²
  Cross-validation: k-fold, stratified k-fold
COMMON WORKFLOW (scikit-learn):
  from sklearn.ensemble import RandomForestClassifier
  from sklearn.model_selection import train_test_split, cross_val_score
  from sklearn.preprocessing import StandardScaler
  X_train,X_test,y_train,y_test = train_test_split(X,y,test_size=0.2)
  model = RandomForestClassifier(); model.fit(X_train,y_train)
  print(model.score(X_test,y_test))
DEEP LEARNING (PyTorch):
  import torch, torch.nn as nn
  class Net(nn.Module):
      def __init__(self): super().__init__(); self.fc=nn.Linear(10,1)
      def forward(self,x): return self.fc(x)
  optimizer=torch.optim.Adam(net.parameters()); loss_fn=nn.MSELoss()
FEATURE ENGINEERING: one-hot encoding, label encoding, normalization (0-1), standardization (z-score),
  missing value imputation, outlier removal, log transforms, polynomial features
OVERFITTING FIXES: regularization (L1/L2), dropout, early stopping, data augmentation, cross-validation

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION 8: DATA STRUCTURES & ALGORITHMS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DS: Array O(1) access, LinkedList O(1) insert/delete, Stack (LIFO), Queue (FIFO),
    HashMap O(1) avg, BST O(log n), Heap (priority queue), Graph (adj list/matrix)
SORTING: Bubble O(n²), Merge O(n log n), Quick O(n log n) avg, Heap O(n log n), Counting O(n+k)
SEARCHING: Linear O(n), Binary O(log n) [sorted array], BFS/DFS O(V+E) [graphs]
PATTERNS: Two pointers, Sliding window, Fast/slow pointers, Prefix sum, Monotonic stack
DP: memoization (top-down), tabulation (bottom-up) — Fibonacci, LCS, Knapsack, Coin change
GRAPHS: BFS (shortest path unweighted), DFS (cycle detection, topological sort),
        Dijkstra (weighted shortest path), Union-Find (connected components)
BIG-O: O(1)<O(log n)<O(n)<O(n log n)<O(n²)<O(2ⁿ)<O(n!)
"""


# ─── Universal File Reader ────────────────────────────────────────────────────

def _read_file_content(path: str) -> tuple[str, str]:
    """
    Read any file and return (display_name, extracted_text).
    For code files: numbered lines + metadata (total lines, functions, classes, imports).
    For documents: full text with page/section info.
    Never crashes — always returns something useful.
    """
    p = Path(path)
    ext = p.suffix.lower()
    size = p.stat().st_size if p.exists() else 0

    # ── Plain text / code ────────────────────────────────────────────────
    TEXT_EXTS = {
        '.txt', '.md', '.markdown', '.rst', '.log', '.csv', '.tsv',
        '.py', '.js', '.ts', '.jsx', '.tsx', '.html', '.htm', '.css',
        '.json', '.yaml', '.yml', '.toml', '.ini', '.cfg', '.conf',
        '.xml', '.sql', '.sh', '.bash', '.zsh', '.fish', '.ps1',
        '.bat', '.cmd', '.c', '.cpp', '.cc', '.h', '.hpp', '.cs',
        '.java', '.kt', '.swift', '.go', '.rs', '.rb', '.php',
        '.r', '.m', '.scala', '.dart', '.lua', '.pl', '.ex', '.exs',
        '.vue', '.svelte', '.env', '.gitignore', '.dockerfile',
        '.makefile', '.gradle', '.cmake',
    }

    CODE_EXTS = {
        '.py', '.js', '.ts', '.jsx', '.tsx', '.css', '.html', '.htm',
        '.c', '.cpp', '.cc', '.h', '.hpp', '.cs', '.java', '.kt',
        '.swift', '.go', '.rs', '.rb', '.php', '.r', '.m', '.scala',
        '.dart', '.lua', '.pl', '.ex', '.exs', '.vue', '.svelte',
        '.sh', '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd',
        '.sql', '.xml', '.json', '.yaml', '.yml', '.toml',
    }

    if ext in TEXT_EXTS or ext == '':
        try:
            raw = p.read_text(encoding='utf-8', errors='replace')
            lines = raw.splitlines()
            total_lines = len(lines)
            lang = ext.lstrip('.') or 'text'

            # ── Add line numbers ──────────────────────────────────────────
            numbered_lines = []
            for i, line in enumerate(lines, 1):
                numbered_lines.append(f"{i:>6} | {line}")
            numbered_text = "\n".join(numbered_lines)

            # ── Build metadata block ──────────────────────────────────────
            meta_lines = [
                f"FILE METADATA",
                f"  Name       : {p.name}",
                f"  Size       : {size:,} bytes",
                f"  Total lines: {total_lines}",
                f"  Language   : {lang}",
            ]

            # Code-specific analysis
            if ext in CODE_EXTS:
                blank_lines = sum(1 for l in lines if not l.strip())
                comment_lines = 0
                functions = []
                classes = []
                imports = []

                if ext == '.py':
                    for i, line in enumerate(lines, 1):
                        s = line.strip()
                        if s.startswith('#') or s.startswith('"""') or s.startswith("'''"):
                            comment_lines += 1
                        if s.startswith('def ') or s.startswith('async def '):
                            fname = s.split('(')[0].replace('def ', '').replace('async ', '').strip()
                            functions.append(f"line {i}: {fname}")
                        if s.startswith('class '):
                            cname = s.split('(')[0].split(':')[0].replace('class ', '').strip()
                            classes.append(f"line {i}: {cname}")
                        if s.startswith('import ') or s.startswith('from '):
                            imports.append(f"line {i}: {s[:80]}")
                elif ext in ('.js', '.ts', '.jsx', '.tsx'):
                    for i, line in enumerate(lines, 1):
                        s = line.strip()
                        if s.startswith('//') or s.startswith('/*') or s.startswith('*'):
                            comment_lines += 1
                        if 'function ' in s or s.startswith('const ') and '=>' in s:
                            functions.append(f"line {i}: {s[:60]}")
                        if s.startswith('class '):
                            cname = s.split('{')[0].split('extends')[0].replace('class ', '').strip()
                            classes.append(f"line {i}: {cname}")
                        if s.startswith('import ') or s.startswith('require('):
                            imports.append(f"line {i}: {s[:80]}")
                elif ext in ('.java', '.cs', '.kt', '.swift'):
                    for i, line in enumerate(lines, 1):
                        s = line.strip()
                        if s.startswith('//') or s.startswith('/*') or s.startswith('*'):
                            comment_lines += 1
                        if 'void ' in s or 'public ' in s or 'private ' in s or 'protected ' in s:
                            if '(' in s and ')' in s and '{' in s:
                                functions.append(f"line {i}: {s[:60]}")
                        if s.startswith('class ') or 'class ' in s:
                            classes.append(f"line {i}: {s[:60]}")
                        if s.startswith('import ') or s.startswith('using ') or s.startswith('package '):
                            imports.append(f"line {i}: {s[:80]}")
                else:
                    for i, line in enumerate(lines, 1):
                        s = line.strip()
                        if s.startswith('#') or s.startswith('//') or s.startswith('/*'):
                            comment_lines += 1

                meta_lines += [
                    f"  Code lines : {total_lines - blank_lines - comment_lines}",
                    f"  Blank lines: {blank_lines}",
                    f"  Comments   : {comment_lines}",
                ]
                if classes:
                    meta_lines.append(f"\nCLASSES ({len(classes)}):")
                    meta_lines += [f"    {c}" for c in classes[:50]]
                if functions:
                    meta_lines.append(f"\nFUNCTIONS / METHODS ({len(functions)}):")
                    meta_lines += [f"    {f}" for f in functions[:100]]
                if imports:
                    meta_lines.append(f"\nIMPORTS ({len(imports)}):")
                    meta_lines += [f"    {im}" for im in imports[:50]]

            # For plain text / docs
            elif ext in ('.txt', '.md', '.markdown', '.rst', '.log'):
                words = len(raw.split())
                meta_lines.append(f"  Word count : {words:,}")
                # Headings for markdown
                if ext in ('.md', '.markdown', '.rst'):
                    headings = [(i, l) for i, l in enumerate(lines, 1) if l.startswith('#') or (len(l) > 1 and l[1:].startswith('==') or l[1:].startswith('--'))]
                    if headings:
                        meta_lines.append(f"\nHEADINGS / SECTIONS:")
                        for ln, h in headings[:30]:
                            meta_lines.append(f"    line {ln}: {h[:80]}")

            meta_block = "\n".join(meta_lines)

            return p.name, (
                f"=== FILE: {p.name} ===\n"
                f"{meta_block}\n\n"
                f"=== FULL CONTENT (with line numbers) ===\n"
                f"```{lang}\n{numbered_text}\n```"
            )
        except Exception as e:
            return p.name, f"[Could not read {p.name}: {e}]"

    # ── PDF ──────────────────────────────────────────────────────────────
    if ext == '.pdf':
        try:
            try:
                from pypdf import PdfReader
            except ImportError:
                try:
                    from PyPDF2 import PdfReader
                except ImportError:
                    PdfReader = None
            if PdfReader:
                reader = PdfReader(str(p))
                total_pages = len(reader.pages)
                pages = []
                all_text = []
                for i, page in enumerate(reader.pages[:100]):
                    page_text = page.extract_text() or ''
                    all_text.append(page_text)
                    pages.append(f"--- Page {i+1} of {total_pages} ---\n{page_text}")
                full_text = "\n".join(pages)
                total_words = len(" ".join(all_text).split())
                total_chars = sum(len(t) for t in all_text)
                meta = (
                    f"FILE METADATA\n"
                    f"  Name       : {p.name}\n"
                    f"  Size       : {size:,} bytes\n"
                    f"  Pages      : {total_pages}\n"
                    f"  Words      : {total_words:,}\n"
                    f"  Characters : {total_chars:,}\n"
                )
                return p.name, (
                    f"=== FILE: {p.name} ===\n{meta}\n"
                    f"=== FULL CONTENT ===\n{full_text}"
                )
            return p.name, f"[PDF: {p.name} — install pypdf: pip install pypdf]"
        except Exception as e:
            return p.name, f"[PDF read error: {p.name}: {e}]"

    # ── DOCX ─────────────────────────────────────────────────────────────
    if ext == '.docx':
        try:
            try:
                from docx import Document
                doc = Document(str(p))
                paragraphs = [(i+1, para.style.name, para.text)
                              for i, para in enumerate(doc.paragraphs) if para.text.strip()]
                headings = [(i, style, txt) for i, style, txt in paragraphs
                            if 'heading' in style.lower() or 'title' in style.lower()]
                full_text = "\n\n".join(f"[Para {i}] {txt}" for i, _, txt in paragraphs)
                words = len(" ".join(p[2] for p in paragraphs).split())
                meta = (
                    f"FILE METADATA\n"
                    f"  Name       : {p.name}\n"
                    f"  Size       : {size:,} bytes\n"
                    f"  Paragraphs : {len(paragraphs)}\n"
                    f"  Words      : {words:,}\n"
                )
                structure = ""
                if headings:
                    structure = "\nDOCUMENT STRUCTURE (headings):\n" + "\n".join(
                        f"  Para {i}: [{style}] {txt[:100]}" for i, style, txt in headings[:40]
                    )
                return p.name, (
                    f"=== FILE: {p.name} ===\n{meta}{structure}\n\n"
                    f"=== FULL CONTENT ===\n{full_text}"
                )
            except ImportError:
                import zipfile, xml.etree.ElementTree as ET
                with zipfile.ZipFile(str(p)) as z:
                    if 'word/document.xml' in z.namelist():
                        xml_content = z.read('word/document.xml').decode('utf-8', errors='replace')
                        root = ET.fromstring(xml_content)
                        texts = [node.text or '' for node in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t')]
                        text = ' '.join(t for t in texts if t.strip())
                        return p.name, f"=== FILE: {p.name} ===\n[Word Document — install python-docx for richer extraction]\n\n{text}"
                return p.name, f"[DOCX: {p.name} — install python-docx: pip install python-docx]"
        except Exception as e:
            return p.name, f"[DOCX read error: {p.name}: {e}]"

    # ── XLSX / XLS ───────────────────────────────────────────────────────
    if ext in ('.xlsx', '.xls', '.xlsm'):
        try:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), data_only=True)
                sheets = []
                total_rows = 0
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_vals = [str(c) if c is not None else '' for c in row]
                        if any(v.strip() for v in row_vals):
                            rows.append('\t'.join(row_vals))
                    total_rows += len(rows)
                    if rows:
                        sheets.append(
                            f"=== Sheet: {sheet_name} ({len(rows)} rows) ===\n"
                            + '\n'.join(rows[:500])
                        )
                text = '\n\n'.join(sheets)
                meta = (
                    f"FILE METADATA\n"
                    f"  Name   : {p.name}\n"
                    f"  Size   : {size:,} bytes\n"
                    f"  Sheets : {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})\n"
                    f"  Rows   : {total_rows:,} total (capped at 500/sheet)\n"
                )
                # ── Auto-ingest into Excel Intelligence DB ──────────────────
                if _excel_intel:
                    try:
                        ingest_result = _excel_intel.ingest_workbook(str(p))
                        sheets_info = ingest_result.get("sheets", [])
                        for sh in sheets_info:
                            meta += (
                                f"\n  ✓ DB stored: sheet '{sh['sheet']}' "
                                f"({sh['rows']} rows, cols: {', '.join(sh['columns'][:8])})"
                            )
                    except Exception as _ei:
                        meta += f"\n  [Excel DB ingest: {_ei}]"
                # ── End Excel Intelligence ingest ────────────────────────────────
                return p.name, f"=== FILE: {p.name} ===\n{meta}\n=== CONTENT ===\n{text}"
            except ImportError:
                return p.name, f"[Excel: {p.name} — install openpyxl: pip install openpyxl]"
        except Exception as e:
            return p.name, f"[Excel read error: {p.name}: {e}]"

    # ── Images ───────────────────────────────────────────────────────────
    IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg', '.ico', '.tiff', '.tif'}
    if ext in IMAGE_EXTS:
        try:
            mime = mimetypes.guess_type(str(p))[0] or 'image/png'
            if ext == '.svg':
                svg_text = p.read_text(encoding='utf-8', errors='replace')
                return p.name, f"=== FILE: {p.name} ===\n[SVG Image, {size:,} bytes]\n```svg\n{svg_text[:3000]}\n```"
            return p.name, (
                f"=== FILE: {p.name} ===\n"
                f"[Image file: {mime}, {size:,} bytes — binary content not shown as text]"
            )
        except Exception as e:
            return p.name, f"[Image read error: {p.name}: {e}]"

    # ── ZIP / Archives ───────────────────────────────────────────────────
    if ext in ('.zip', '.tar', '.gz', '.tgz'):
        try:
            if ext == '.zip':
                with zipfile.ZipFile(str(p)) as z:
                    names = z.namelist()
                    listing = '\n'.join(f"  {n}" for n in names[:200])
                    return p.name, (
                        f"=== FILE: {p.name} ===\n"
                        f"[ZIP Archive, {size:,} bytes, {len(names)} files]\n\n"
                        f"CONTENTS:\n{listing}"
                    )
            return p.name, f"=== FILE: {p.name} ===\n[Archive ({ext}), {size:,} bytes]"
        except Exception as e:
            return p.name, f"[Archive read error: {p.name}: {e}]"

    # ── Fallback: binary / unknown ───────────────────────────────────────
    try:
        raw = p.read_text(encoding='utf-8', errors='replace')
        lines = raw.splitlines()
        numbered = "\n".join(f"{i:>6} | {l}" for i, l in enumerate(lines, 1))
        return p.name, (
            f"=== FILE: {p.name} ===\n"
            f"FILE METADATA\n  Name: {p.name}\n  Size: {size:,} bytes\n  Lines: {len(lines)}\n\n"
            f"=== FULL CONTENT (with line numbers) ===\n{numbered}"
        )
    except Exception:
        return p.name, f"[Binary file: {p.name} ({ext}, {size:,} bytes) — not readable as text]"


def _file_icon(path: str) -> str:
    """Return an emoji icon for a file based on its extension."""
    ext = Path(path).suffix.lower()
    icons = {
        '.py': '🐍', '.js': '🟨', '.ts': '🔷', '.jsx': '⚛', '.tsx': '⚛',
        '.html': '🌐', '.css': '🎨', '.json': '📋', '.yaml': '📋', '.yml': '📋',
        '.md': '📝', '.txt': '📄', '.pdf': '📕', '.docx': '📘', '.doc': '📘',
        '.xlsx': '📗', '.xls': '📗', '.csv': '📊', '.png': '🖼', '.jpg': '🖼',
        '.jpeg': '🖼', '.gif': '🖼', '.svg': '🎨', '.zip': '📦', '.tar': '📦',
        '.gz': '📦', '.sh': '⚙', '.bat': '⚙', '.sql': '🗃', '.xml': '📋',
        '.java': '☕', '.cpp': '⚙', '.c': '⚙', '.rs': '🦀', '.go': '🐹',
        '.rb': '💎', '.php': '🐘', '.swift': '🍎', '.kt': '🔶',
    }
    return icons.get(ext, '📎')

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTextEdit, QLineEdit, QLabel, QComboBox,
    QScrollArea, QFrame, QSizePolicy, QGraphicsDropShadowEffect,
    QGraphicsOpacityEffect,
    QSpacerItem, QProgressBar, QListWidget, QListWidgetItem,
    QFileDialog, QMessageBox, QInputDialog, QCheckBox, QDialog,
    QDialogButtonBox, QTabWidget, QSplitter, QStackedWidget,
    QSystemTrayIcon, QMenu, QSizeGrip
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QTimer, QPropertyAnimation,
    QEasingCurve, QSize, QPoint, QRect, QObject,
    QParallelAnimationGroup, QSequentialAnimationGroup,
    QRectF, QPointF
)
from PySide6.QtGui import (
    QFont, QColor, QPalette, QLinearGradient, QPainter,
    QBrush, QPen, QPixmap, QFontDatabase, QTextCursor,
    QIcon, QKeySequence, QShortcut, QClipboard,
    QRadialGradient, QPainterPath, QPolygon, QAction
)


# ─── Multilingual System Prompts ──────────────────────────────────────────────

BASE_PERSONALITY = {
    "formal": (
        "You are Digi Valet, a refined and highly capable personal AI assistant. "
        "Your personality is professional, articulate, and discreet — like a trusted personal valet "
        "who anticipates needs before they are spoken. You help with tasks like scheduling, writing, "
        "research, advice, planning, and any personal or professional matter the user brings to you.\n\n"
        "Guidelines:\n"
        "- Be precise and thorough. Use formal, elegant language.\n"
        "- Avoid colloquialisms or casual phrasing.\n"
        "- Remember context within this conversation.\n"
        "- Never break character. You are Digi Valet.\n"
        "- If you don't know something, say so gracefully and offer to help find out.\n"
    ),
    "balanced": (
        "You are Digi Valet, a capable and friendly personal AI assistant. "
        "You're warm, helpful, and professional — striking a balance between polished and approachable. "
        "You help with tasks like scheduling, writing, research, advice, planning, and everyday needs.\n\n"
        "Guidelines:\n"
        "- Be clear, concise, and helpful.\n"
        "- Use natural language — not too stiff, not too casual.\n"
        "- Remember context within this conversation.\n"
        "- Never break character. You are Digi Valet.\n"
        "- If you don't know something, say so and offer to help.\n"
    ),
    "casual": (
        "You are Digi Valet, a smart and easygoing personal AI assistant. "
        "You're friendly, conversational, and helpful — like a knowledgeable friend who's always ready to help.\n\n"
        "Guidelines:\n"
        "- Be direct and approachable. Feel free to use casual language and contractions.\n"
        "- Keep things light but accurate.\n"
        "- Remember context within this conversation.\n"
        "- You're still Digi Valet — just a relaxed version.\n"
        "- If you're not sure about something, be upfront and offer to find out.\n"
    ),
}

LANGUAGE_ADDONS = {
    "English": "",
    "Hindi": "Always respond in Hindi (Devanagari script). The user prefers to communicate in Hindi.\n",
    "Spanish": "Always respond in Spanish. The user prefers to communicate in Spanish.\n",
    "French": "Always respond in French. The user prefers to communicate in French.\n",
    "German": "Always respond in German. The user prefers to communicate in German.\n",
    "Arabic": "Always respond in Arabic. The user prefers to communicate in Arabic.\n",
    "Japanese": "Always respond in Japanese. The user prefers to communicate in Japanese.\n",
    "Portuguese": "Always respond in Portuguese. The user prefers to communicate in Portuguese.\n",
}

QUICK_COMMANDS = {
    "/help": (
        "List all available quick-commands I can use, and give me a brief summary "
        "of your capabilities as Digi Valet."
    ),
    "/tasks": (
        "Show me my current task list and ask me what I'd like to add, complete, or remove."
    ),
    "/plan": (
        "Help me plan my day. Ask me what's on my agenda and help me prioritize and schedule it."
    ),
    "/wellness": (
        "Give me a brief wellness check-in. Ask me about my mood, energy, hydration, and sleep, "
        "then give me a personalized tip."
    ),
    "/meal": (
        "Suggest a healthy, balanced meal plan for today. Ask me about any dietary restrictions first."
    ),
    "/summarize": (
        "Summarize our conversation so far in bullet points."
    ),
    "/focus": (
        "Help me set a focus session. Ask me what I'm working on and help me set a goal and timer strategy."
    ),
    "/digivalet": (
        "Show me how to use the /digivalet org-chart lookup command "
        "(e.g. /digivalet Rahul Salgia, /digivalet QA, /digivalet stats)."
    ),
    "/kb": "/kb list",  # handled locally — shows KB status directly
    "/excel": ("Show me a summary of all Excel files loaded into your database, "
               "including sheet names, row counts, and column names. "
               "Tell me what graphs I can make from this data."),
}


def build_system_prompt(tone: str = "balanced", language: str = "English",
                        kb=None, include_programming_kb: bool = True,
                        org_tree=None, excel_intel=None) -> str:
    """Build the full Ollama system prompt with all knowledge sources merged."""
    base = BASE_PERSONALITY.get(tone, BASE_PERSONALITY["balanced"])
    addon = LANGUAGE_ADDONS.get(language, "")
    prompt = base + ("\n" + addon if addon else "")
    if kb is not None and not kb.is_empty():
        prompt += "\n\n" + kb.system_prompt_block()
    if org_tree is not None and org_tree.loaded:
        snippet = org_tree.to_context_snippet(max_rows=200)
        if snippet:
            prompt += "\n\n# Employee Directory (answer org-chart questions from this)\n" + snippet
    if excel_intel is not None and excel_intel.has_data():
        prompt += excel_intel.build_excel_context()
    if PROGRAMMING_KNOWLEDGE and include_programming_kb:
        prompt += "\n\n" + PROGRAMMING_KNOWLEDGE
    return prompt


_CODE_FENCE_RE = re.compile(r"```")


def _message_likely_contains_code(text: str) -> bool:
    """Heuristic: does this user message carry an actual code paste or
    file attachment that the model should focus on, rather than a
    generic 'how do I do X' syntax question?"""
    if not text:
        return False
    if _CODE_FENCE_RE.search(text):
        return True
    if "**Attached files for context:**" in text:
        return True
    # A handful of lines with classic code punctuation/indentation is a
    # decent signal even without fences.
    code_hint_lines = sum(
        1 for line in text.splitlines()
        if line.strip() and (line.startswith((" ", "\t"))
                              or any(tok in line for tok in ("def ", "class ", "import ", "{", "}", ";")))
    )
    return code_hint_lines >= 3


# ─── Ollama Worker Thread ──────────────────────────────────────────────────────

class OllamaWorker(QObject):
    token_received = Signal(str)
    finished = Signal()
    error = Signal(str)

    def __init__(self, model: str, messages: list, host: str = "http://localhost:11434"):
        super().__init__()
        self.model = model
        self.messages = messages
        self.host = host
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        payload = json.dumps({
            "model": self.model,
            "messages": self.messages,
            "stream": True,
            # Ollama defaults to a tiny 2048-token context unless told
            # otherwise. With a large system prompt (knowledge base) plus
            # pasted/attached code, that default silently truncates the
            # user's actual code out of the context — which is why the
            # model can "forget" code that was just sent. Force a much
            # larger window so real-world code pastes survive.
            "options": {"num_ctx": 8192}
        }).encode()

        try:
            req = urllib.request.Request(
                f"{self.host}/api/chat",
                data=payload,
                headers={"Content-Type": "application/json"},
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                for raw_line in resp:
                    if self._cancelled:
                        break
                    line = raw_line.decode("utf-8").strip()
                    if not line:
                        continue
                    try:
                        data = json.loads(line)
                        content = data.get("message", {}).get("content", "")
                        if content:
                            self.token_received.emit(content)
                        if data.get("done"):
                            break
                    except json.JSONDecodeError:
                        pass
        except urllib.error.URLError as e:
            if not self._cancelled:
                self.error.emit(
                    f"Cannot connect to Ollama.\n\nMake sure Ollama is running:\n  ollama serve\n\nError: {e.reason}"
                )
        except Exception as e:
            if not self._cancelled:
                self.error.emit(str(e))
        finally:
            self.finished.emit()


class OllamaModelsWorker(QObject):
    models_ready = Signal(list)
    error = Signal(str)

    def __init__(self, host="http://localhost:11434"):
        super().__init__()
        self.host = host

    def run(self):
        try:
            req = urllib.request.Request(f"{self.host}/api/tags")
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read())
                models = [m["name"] for m in data.get("models", [])]
                self.models_ready.emit(models)
        except Exception as e:
            self.error.emit(str(e))


# ─── Markdown Renderer ────────────────────────────────────────────────────────

def markdown_to_html(text: str, dark_mode: bool = True) -> str:
    """Convert a subset of Markdown to HTML for QLabel rich text display."""
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    code_bg = "#1a1a12" if dark_mode else "#f0ede6"
    code_fg = "#b8d0a0" if dark_mode else "#3a6040"
    header_color = "#c8a96e" if dark_mode else "#8b6914"
    quote_fg = "#8a7a6a" if dark_mode else "#6a5a4a"

    def replace_code_block(m):
        lang = m.group(1).strip()
        code = m.group(2).strip()
        # Unescape HTML entities inside code blocks so code renders correctly
        code = code.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
        lang_label = f'<span style="color:{header_color};font-size:9px;opacity:0.7;">{lang}</span><br>' if lang else ""
        return (
            f'<div style="background:{code_bg};border:1px solid #3a3028;border-left:3px solid {header_color};'
            f'border-radius:6px;padding:8px 12px;margin:6px 0;font-family:monospace;font-size:11px;'
            f'color:{code_fg};white-space:pre-wrap;">{lang_label}{code}</div>'
        )
    text = re.sub(r'```(\w*)\n(.*?)```', replace_code_block, text, flags=re.DOTALL)

    text = re.sub(
        r'`([^`]+)`',
        rf'<code style="background:{code_bg};color:{code_fg};padding:1px 5px;border-radius:3px;font-family:monospace;font-size:11px;">\1</code>',
        text
    )

    text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
    text = re.sub(r'__(.+?)__', r'<b>\1</b>', text)
    text = re.sub(r'\*([^*\n]+)\*', r'<i>\1</i>', text)
    text = re.sub(r'_([^_\n]+)_', r'<i>\1</i>', text)
    text = re.sub(r'~~(.+?)~~', r'<s>\1</s>', text)

    text = re.sub(r'^### (.+)$', rf'<span style="font-size:13px;color:{header_color};font-weight:bold;">\1</span>', text, flags=re.MULTILINE)
    text = re.sub(r'^## (.+)$',  rf'<span style="font-size:14px;color:{header_color};font-weight:bold;">\1</span>', text, flags=re.MULTILINE)
    text = re.sub(r'^# (.+)$',   rf'<span style="font-size:15px;color:{header_color};font-weight:bold;">\1</span>', text, flags=re.MULTILINE)

    text = re.sub(r'^-{3,}$', f'<hr style="border:none;border-top:1px solid #3a3028;margin:8px 0;"/>', text, flags=re.MULTILINE)

    def replace_ul(m):
        items = re.findall(r'^[\-\*] (.+)$', m.group(0), re.MULTILINE)
        li_html = ''.join(f'<li style="margin:2px 0;">{i}</li>' for i in items)
        return f'<ul style="margin:4px 0 4px 16px;padding:0;">{li_html}</ul>'
    text = re.sub(r'(^[\-\*] .+\n?)+', replace_ul, text, flags=re.MULTILINE)

    def replace_ol(m):
        items = re.findall(r'^\d+\. (.+)$', m.group(0), re.MULTILINE)
        li_html = ''.join(f'<li style="margin:2px 0;">{i}</li>' for i in items)
        return f'<ol style="margin:4px 0 4px 16px;padding:0;">{li_html}</ol>'
    text = re.sub(r'(^\d+\. .+\n?)+', replace_ol, text, flags=re.MULTILINE)

    text = re.sub(
        r'^&gt; (.+)$',
        rf'<div style="border-left:3px solid {header_color};padding:4px 10px;color:{quote_fg};margin:4px 0;">\1</div>',
        text, flags=re.MULTILINE
    )

    text = text.replace('\n', '<br>')
    return text


# ─── Message Bubble Widget ─────────────────────────────────────────────────────

_FONT_SIZE_DELTA = 0
_DARK_MODE = True


class MessageBubble(QFrame):
    def __init__(self, text: str, role: str, parent=None):
        super().__init__(parent)
        self.role = role
        self._raw_text = text
        self.setObjectName("bubble")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(4)

        header = QHBoxLayout()
        header.setSpacing(6)

        icon_label = QLabel()
        icon_label.setFixedSize(22, 22)
        icon_label.setAlignment(Qt.AlignCenter)

        name_label = QLabel()
        name_label.setFont(QFont("Georgia", 9))

        time_label = QLabel(datetime.now().strftime("%H:%M"))
        time_label.setFont(QFont("Georgia", 8))
        time_label.setObjectName("timeLabel")

        self.copy_btn = QPushButton("⎘")
        self.copy_btn.setObjectName("copyBtn")
        self.copy_btn.setFixedSize(22, 22)
        self.copy_btn.setToolTip("Copy message")
        self.copy_btn.setFont(QFont("Georgia", 10))
        self.copy_btn.clicked.connect(self._copy_text)
        self.copy_btn.setCursor(Qt.PointingHandCursor)

        if role == "assistant":
            icon_label.setText("◈")
            icon_label.setStyleSheet("color: #c8a96e; font-size: 14px;")
            name_label.setText("Digi Valet")
            name_label.setStyleSheet("color: #c8a96e; font-weight: bold;")
            self.setObjectName("assistantBubble")
        else:
            icon_label.setText("◉")
            icon_label.setStyleSheet("color: #8ab4c8; font-size: 13px;")
            name_label.setText("You")
            name_label.setStyleSheet("color: #8ab4c8; font-weight: bold;")
            self.setObjectName("userBubble")

        header.addWidget(icon_label)
        header.addWidget(name_label)
        header.addStretch()
        header.addWidget(time_label)
        header.addSpacing(6)
        header.addWidget(self.copy_btn)
        layout.addLayout(header)

        self.text_label = QLabel()
        self.text_label.setWordWrap(True)
        self.text_label.setTextInteractionFlags(
            Qt.TextSelectableByMouse | Qt.LinksAccessibleByMouse
        )
        self.text_label.setOpenExternalLinks(True)
        self._base_font_size = 11 + _FONT_SIZE_DELTA
        self.text_label.setFont(QFont("Georgia", self._base_font_size))
        self.text_label.setObjectName("msgText")
        self.text_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        layout.addWidget(self.text_label)

        self._render(text)

    def _render(self, text: str):
        if text:
            html = markdown_to_html(text, _DARK_MODE)
            self.text_label.setText(html)
        else:
            self.text_label.setText("")

    def _copy_text(self):
        QApplication.clipboard().setText(self._raw_text)
        self.copy_btn.setText("✓")
        QTimer.singleShot(1200, lambda: self.copy_btn.setText("⎘") if not self.copy_btn.isHidden() else None)

    def append_text(self, chunk: str):
        self._raw_text += chunk
        self._render(self._raw_text)

    def set_text(self, text: str):
        self._raw_text = text
        self._render(text)

    def update_font_size(self, delta: int):
        self._base_font_size = max(8, min(20, 11 + delta))
        self.text_label.setFont(QFont("Georgia", self._base_font_size))


# ─── Typing Indicator ─────────────────────────────────────────────────────────

class TypingIndicator(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("typingIndicator")
        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(8)

        icon = QLabel("◈")
        icon.setStyleSheet("color: #c8a96e; font-size: 14px;")
        layout.addWidget(icon)

        self.dot_label = QLabel("Digi Valet is thinking")
        self.dot_label.setFont(QFont("Georgia", 10))
        self.dot_label.setStyleSheet("color: #c8a96e;")
        layout.addWidget(self.dot_label)
        layout.addStretch()

        self._dots = 0
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._animate)
        self._timer.start(500)

    def _animate(self):
        self._dots = (self._dots + 1) % 4
        self.dot_label.setText("Digi Valet is thinking" + "." * self._dots)

    def stop(self):
        if self._timer and self._timer.isActive():
            self._timer.stop()


# ─── Thinking Step Card (Claude-style collapsible process steps) ──────────────

class ThinkingStepCard(QFrame):
    """
    A collapsible card that shows a single AI processing step.
    Click the header to expand/collapse the detail.
    """
    def __init__(self, step_num: int, title: str, detail: str = "", parent=None):
        super().__init__(parent)
        self.setObjectName("stepCard")
        self._expanded = False
        self._detail = detail

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ── Header row (always visible) ──────────────────────────────────
        self.header = QPushButton()
        self.header.setObjectName("stepHeader")
        self.header.setCursor(Qt.PointingHandCursor)
        self.header.setCheckable(False)
        self.header.clicked.connect(self._toggle)

        h_layout = QHBoxLayout(self.header)
        h_layout.setContentsMargins(10, 7, 10, 7)
        h_layout.setSpacing(8)

        self.num_badge = QLabel(str(step_num))
        self.num_badge.setObjectName("stepBadge")
        self.num_badge.setFixedSize(20, 20)
        self.num_badge.setAlignment(Qt.AlignCenter)
        self.num_badge.setFont(QFont("Georgia", 8, QFont.Bold))
        h_layout.addWidget(self.num_badge)

        self.title_lbl = QLabel(title)
        self.title_lbl.setObjectName("stepTitle")
        self.title_lbl.setFont(QFont("Georgia", 10))
        h_layout.addWidget(self.title_lbl, 1)

        self.chevron = QLabel("▸")
        self.chevron.setObjectName("stepChevron")
        self.chevron.setFont(QFont("Georgia", 10))
        h_layout.addWidget(self.chevron)

        layout.addWidget(self.header)

        # ── Detail area (hidden until expanded) ─────────────────────────
        self.detail_frame = QFrame()
        self.detail_frame.setObjectName("stepDetail")
        d_layout = QVBoxLayout(self.detail_frame)
        d_layout.setContentsMargins(38, 6, 12, 10)

        self.detail_lbl = QLabel(detail or "Processing…")
        self.detail_lbl.setObjectName("stepDetailText")
        self.detail_lbl.setWordWrap(True)
        self.detail_lbl.setFont(QFont("Georgia", 9))
        self.detail_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        d_layout.addWidget(self.detail_lbl)

        self.detail_frame.hide()
        layout.addWidget(self.detail_frame)

        self._apply_style()

    def _apply_style(self):
        dark = _DARK_MODE
        if dark:
            border = "#2a2420"
            badge_bg = "#c8a96e"
            badge_fg = "#0e0c0a"
            title_fg = "#d4cfc8"
            chevron_fg = "#8a7a6a"
            detail_bg = "#141210"
            detail_fg = "#a09888"
            header_bg = "#1a1612"
            header_hover = "#221e1a"
        else:
            border = "#d8d0c0"
            badge_bg = "#c8a96e"
            badge_fg = "#fff8ec"
            title_fg = "#2a2010"
            chevron_fg = "#a09070"
            detail_bg = "#fffdf5"
            detail_fg = "#706050"
            header_bg = "#f5f0e8"
            header_hover = "#eee8d8"

        self.setStyleSheet(f"""
            QFrame#stepCard {{ border: 1px solid {border}; border-radius: 8px;
                               margin: 2px 0px; background: transparent; }}
            QPushButton#stepHeader {{ background: {header_bg}; border: none;
                                       border-radius: 8px; text-align: left; }}
            QPushButton#stepHeader:hover {{ background: {header_hover}; }}
            QLabel#stepBadge {{ background: {badge_bg}; color: {badge_fg};
                                 border-radius: 10px; font-weight: bold; }}
            QLabel#stepTitle {{ color: {title_fg}; background: transparent; }}
            QLabel#stepChevron {{ color: {chevron_fg}; background: transparent; }}
            QFrame#stepDetail {{ background: {detail_bg}; border-top: 1px solid {border};
                                  border-bottom-left-radius: 8px;
                                  border-bottom-right-radius: 8px; }}
            QLabel#stepDetailText {{ color: {detail_fg}; background: transparent; }}
        """)

    def set_detail(self, detail: str):
        self._detail = detail
        self.detail_lbl.setText(detail)

    def set_done(self, success: bool = True):
        self.num_badge.setText("✓" if success else "✗")
        color = "#6a9e6a" if success else "#e07070"
        self.num_badge.setStyleSheet(
            f"background: {color}; color: white; border-radius: 10px; font-weight: bold;"
        )

    def _toggle(self):
        self._expanded = not self._expanded
        self.chevron.setText("▾" if self._expanded else "▸")
        if self._expanded:
            self.detail_frame.show()
        else:
            self.detail_frame.hide()

    def expand(self):
        if not self._expanded:
            self._toggle()


class ThinkingStepsWidget(QFrame):
    """Container for a sequence of ThinkingStepCards."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("thinkingStepsWidget")
        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(3)
        self._cards = []

        header = QLabel("◈  Processing steps")
        header.setFont(QFont("Georgia", 9))
        header.setObjectName("thinkingHeader")
        self._layout.addWidget(header)

        if _DARK_MODE:
            self.setStyleSheet("QLabel#thinkingHeader { color: #c8a96e; padding: 2px 0 4px 0; background: transparent; }")
        else:
            self.setStyleSheet("QLabel#thinkingHeader { color: #8b6914; padding: 2px 0 4px 0; background: transparent; }")

    def add_step(self, title: str, detail: str = "") -> ThinkingStepCard:
        card = ThinkingStepCard(len(self._cards) + 1, title, detail)
        self._layout.addWidget(card)
        self._cards.append(card)
        return card

    def finish_step(self, card: ThinkingStepCard, detail: str = "", success: bool = True):
        if detail:
            card.set_detail(detail)
        card.set_done(success)


# ─── Code Preview Dialog (Claude-style artifact panel) ────────────────────────

class CodePreviewDialog(QDialog):
    """
    Pops up when a code block is detected in AI response.
    Shows Preview (rendered HTML) and Code tabs side-by-side.
    """
    def __init__(self, code: str, lang: str = "", title: str = "Code Output", parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"◈ Preview — {title}")
        self.resize(900, 600)
        self.setMinimumSize(600, 400)
        self._code = code
        self._lang = lang

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ── Top bar ─────────────────────────────────────────────────────
        top_bar = QWidget()
        top_bar.setFixedHeight(44)
        tbl = QHBoxLayout(top_bar)
        tbl.setContentsMargins(16, 0, 16, 0)
        tbl.setSpacing(8)

        lbl = QLabel(f"◈  {title}")
        lbl.setFont(QFont("Georgia", 11, QFont.Bold))
        tbl.addWidget(lbl)
        tbl.addStretch()

        copy_btn = QPushButton("⎘ Copy code")
        copy_btn.setFont(QFont("Georgia", 9))
        copy_btn.setCursor(Qt.PointingHandCursor)
        copy_btn.clicked.connect(self._copy_code)
        tbl.addWidget(copy_btn)

        close_btn = QPushButton("✕")
        close_btn.setFont(QFont("Georgia", 10))
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.setFixedSize(28, 28)
        close_btn.clicked.connect(self.close)
        tbl.addWidget(close_btn)

        layout.addWidget(top_bar)

        # ── Tab widget ──────────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setFont(QFont("Georgia", 10))
        layout.addWidget(self.tabs)

        # Preview tab
        try:
            from PySide6.QtWebEngineWidgets import QWebEngineView
            self.preview = QWebEngineView()
            html_content = self._build_preview_html(code, lang)
            self.preview.setHtml(html_content)
            self.tabs.addTab(self.preview, "▶  Preview")
        except ImportError:
            no_web = QLabel("Install PySide6-WebEngine for preview:\npip install PySide6-WebEngine")
            no_web.setAlignment(Qt.AlignCenter)
            no_web.setFont(QFont("Georgia", 10))
            self.tabs.addTab(no_web, "▶  Preview")

        # Code tab
        code_widget = QTextEdit()
        code_widget.setReadOnly(True)
        code_widget.setFont(QFont("Courier New", 11))
        code_widget.setPlainText(code)
        if _DARK_MODE:
            code_widget.setStyleSheet(
                "background:#1a1a12; color:#b8d0a0; border:none; padding:12px;"
            )
        else:
            code_widget.setStyleSheet(
                "background:#f8f8f0; color:#2a3a2a; border:none; padding:12px;"
            )
        self.tabs.addTab(code_widget, f"&lt;/&gt;  Code ({lang or 'text'})")

        # Apply styling
        if _DARK_MODE:
            self.setStyleSheet("""
                QDialog { background: #141210; }
                QWidget { background: #141210; color: #d4cfc8; }
                QTabWidget::pane { border: 1px solid #2a2420; border-radius: 0; }
                QTabBar::tab { background: #1a1612; color: #8a7a6a; border: 1px solid #2a2420;
                               padding: 8px 16px; font-family: Georgia; font-size: 10px; }
                QTabBar::tab:selected { background: #141210; color: #c8a96e; border-bottom: 2px solid #c8a96e; }
                QPushButton { background: #1e1a16; color: #c8a96e; border: 1px solid #3a3028;
                              border-radius: 6px; padding: 5px 12px; }
                QPushButton:hover { background: #2a2420; }
            """)
            top_bar.setStyleSheet("background: #0e0c0a; border-bottom: 1px solid #1e1a16;")
            lbl.setStyleSheet("color: #c8a96e;")
        else:
            self.setStyleSheet("""
                QDialog { background: #f5f0e8; }
                QWidget { background: #f5f0e8; color: #2a2010; }
                QTabWidget::pane { border: 1px solid #d0c8b0; }
                QTabBar::tab { background: #ede8dc; color: #706050; border: 1px solid #d0c8b0;
                               padding: 8px 16px; font-family: Georgia; font-size: 10px; }
                QTabBar::tab:selected { background: #f5f0e8; color: #8b6914; border-bottom: 2px solid #c8a96e; }
                QPushButton { background: #fff8ec; color: #8b6914; border: 1px solid #d0b870;
                              border-radius: 6px; padding: 5px 12px; }
                QPushButton:hover { background: #f0e8d0; }
            """)
            top_bar.setStyleSheet("background: #ede8dc; border-bottom: 1px solid #d0c8b0;")
            lbl.setStyleSheet("color: #8b6914;")

        self._copy_btn = copy_btn

    def _copy_code(self):
        QApplication.clipboard().setText(self._code)
        self._copy_btn.setText("✓ Copied!")
        QTimer.singleShot(1500, lambda: self._copy_btn.setText("⎘ Copy code"))

    def _build_preview_html(self, code: str, lang: str) -> str:
        lang = (lang or '').lower()
        bg = "#1a1a12" if _DARK_MODE else "#ffffff"
        fg = "#d4cfc8" if _DARK_MODE else "#2a2010"

        if lang in ('html', 'htm', ''):
            # Try to render as HTML directly
            if '<html' in code.lower() or '<body' in code.lower() or '<div' in code.lower():
                return code
            return f"""<!DOCTYPE html><html><body style="background:{bg};color:{fg};
                       font-family:Georgia,serif;padding:20px;">{code}</body></html>"""
        elif lang in ('markdown', 'md'):
            import html as html_lib
            safe = html_lib.escape(code)
            return f"""<!DOCTYPE html><html><body style="background:{bg};color:{fg};
                       font-family:Georgia,serif;padding:20px;white-space:pre-wrap;">{safe}</body></html>"""
        else:
            import html as html_lib
            safe = html_lib.escape(code)
            return f"""<!DOCTYPE html><html><body style="background:{bg};padding:20px;">
                <pre style="color:#b8d0a0;font-family:monospace;font-size:13px;
                            white-space:pre-wrap;">{safe}</pre></body></html>"""


def _detect_code_blocks(text: str) -> list[tuple[str, str]]:
    """Extract (language, code) pairs from markdown code blocks."""
    pattern = r'```(\w*)\n(.*?)```'
    matches = re.findall(pattern, text, re.DOTALL)
    return [(lang.strip(), code.strip()) for lang, code in matches if code.strip()]


# ─── File Chip Widget ─────────────────────────────────────────────────────────

class FileChip(QFrame):
    """Gemini-style attachment card: icon swatch + filename, with a small
    circular remove button. Fixed compact size so a row of attachments
    stays tidy and doesn't crowd the input bar."""
    removed = Signal(str)  # emits file path

    CARD_WIDTH = 168
    CARD_HEIGHT = 52

    def __init__(self, file_path: str, parent=None):
        super().__init__(parent)
        self._path = file_path
        self.setObjectName("fileChip")
        self.setFixedSize(self.CARD_WIDTH, self.CARD_HEIGHT)

        outer = QHBoxLayout(self)
        outer.setContentsMargins(8, 8, 8, 8)
        outer.setSpacing(8)

        # Icon swatch — small rounded square, like Gemini's file-type tile
        swatch = QLabel(_file_icon(file_path))
        swatch.setObjectName("fileChipSwatch")
        swatch.setFixedSize(36, 36)
        swatch.setAlignment(Qt.AlignCenter)
        swatch.setFont(QFont("Segoe UI Emoji", 14))
        outer.addWidget(swatch)

        # Name + type, stacked
        text_col = QVBoxLayout()
        text_col.setSpacing(1)
        text_col.setContentsMargins(0, 0, 0, 0)

        name = Path(file_path).name
        stem, ext = Path(file_path).stem, Path(file_path).suffix.lstrip(".").upper()
        short_stem = stem if len(stem) <= 16 else stem[:15] + "…"
        name_lbl = QLabel(short_stem)
        name_lbl.setObjectName("fileChipName")
        name_lbl.setFont(QFont("Georgia", 9))
        name_lbl.setToolTip(name)
        text_col.addWidget(name_lbl)

        type_lbl = QLabel(ext or "FILE")
        type_lbl.setObjectName("fileChipType")
        type_lbl.setFont(QFont("Georgia", 8))
        text_col.addWidget(type_lbl)

        outer.addLayout(text_col, 1)

        # Remove button — small circle, top-right
        remove_btn = QPushButton("✕")
        remove_btn.setObjectName("fileChipRemove")
        remove_btn.setFixedSize(18, 18)
        remove_btn.setFont(QFont("Georgia", 8))
        remove_btn.setCursor(Qt.PointingHandCursor)
        remove_btn.setToolTip("Remove")
        remove_btn.clicked.connect(lambda: self.removed.emit(self._path))
        outer.addWidget(remove_btn, 0, Qt.AlignTop)

        self._apply_style()

    def _apply_style(self):
        if _DARK_MODE:
            self.setStyleSheet("""
                QFrame#fileChip { background:#181410; border:1px solid #2a2420;
                                   border-radius:12px; }
                QLabel#fileChipSwatch { background:#241e16; border-radius:8px;
                                          color:#c8a96e; }
                QLabel#fileChipName { color:#d4cfc8; background:transparent; }
                QLabel#fileChipType { color:#5c5046; background:transparent;
                                        letter-spacing:1px; }
                QPushButton#fileChipRemove { background:#241e16; color:#5a4a38;
                                               border:none; border-radius:9px; }
                QPushButton#fileChipRemove:hover { color:#e07070; background:#2a2420; }
            """)
        else:
            self.setStyleSheet("""
                QFrame#fileChip { background:#ffffff; border:1px solid #e0d8c4;
                                   border-radius:12px; }
                QLabel#fileChipSwatch { background:#fff0cc; border-radius:8px;
                                          color:#8b6914; }
                QLabel#fileChipName { color:#2a2010; background:transparent; }
                QLabel#fileChipType { color:#9a8a70; background:transparent;
                                        letter-spacing:1px; }
                QPushButton#fileChipRemove { background:#f0e8d0; color:#a09070;
                                               border:none; border-radius:9px; }
                QPushButton#fileChipRemove:hover { color:#c04040; background:#f8d8d8; }
            """)


# ─── Custom Input Box ─────────────────────────────────────────────────────────

class SmartTextEdit(QTextEdit):
    """Custom QTextEdit: Enter sends, Shift+Enter inserts newline.

    Gemini-style auto-grow: starts as a compact single line and only
    expands as the user types more, capped at MAX_HEIGHT. Resizing is
    debounced with a single-shot QTimer so rapid keystrokes don't trigger
    a relayout on every character — keeps it smooth and lag-free.
    """
    send_requested = Signal()

    MIN_HEIGHT = 44
    MAX_HEIGHT = 160

    def __init__(self, parent=None):
        super().__init__(parent)
        self._resize_timer = QTimer(self)
        self._resize_timer.setSingleShot(True)
        self._resize_timer.setInterval(0)  # coalesce same-frame changes
        self._resize_timer.timeout.connect(self._apply_auto_height)
        self.document().contentsChanged.connect(self._schedule_resize)
        # Lock height immediately so nothing can stretch this box before
        # the first resize pass runs.
        self.setFixedHeight(self.MIN_HEIGHT)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if event.modifiers() & Qt.ShiftModifier:
                super().keyPressEvent(event)
            else:
                self.send_requested.emit()
        else:
            super().keyPressEvent(event)

    def _schedule_resize(self):
        if not self._resize_timer.isActive():
            self._resize_timer.start()

    def _apply_auto_height(self):
        doc_height = int(self.document().size().height())
        if doc_height <= 0:
            doc_height = self.MIN_HEIGHT
        margins = self.contentsMargins()
        target = doc_height + margins.top() + margins.bottom() + 12
        target = max(self.MIN_HEIGHT, min(target, self.MAX_HEIGHT))
        if target != self.height():
            self.setFixedHeight(target)
        self.setVerticalScrollBarPolicy(
            Qt.ScrollBarAsNeeded if doc_height > self.MAX_HEIGHT else Qt.ScrollBarAlwaysOff
        )

    def clear(self):
        super().clear()
        self._apply_auto_height()

    def focusInEvent(self, event):
        super().focusInEvent(event)
        pill = self.parentWidget()
        if pill is not None:
            pill.setProperty("focused", True)
            pill.style().unpolish(pill)
            pill.style().polish(pill)

    def focusOutEvent(self, event):
        super().focusOutEvent(event)
        pill = self.parentWidget()
        if pill is not None:
            pill.setProperty("focused", False)
            pill.style().unpolish(pill)
            pill.style().polish(pill)


# ─── Splash / Welcome Screen ───────────────────────────────────────────────────

def _scaled_rect(rect: QRect, factor: float, focal_point: QPoint) -> QRect:
    """Return `rect` scaled by `factor` while keeping `focal_point` fixed in place.

    Used to produce a "zoom toward the click point" effect: a factor > 1
    grows the rect outward from the focal point (zoom in), a factor < 1
    shrinks it toward the focal point (zoom out)."""
    new_w = rect.width() * factor
    new_h = rect.height() * factor
    new_x = focal_point.x() - (focal_point.x() - rect.x()) * factor
    new_y = focal_point.y() - (focal_point.y() - rect.y()) * factor
    return QRect(int(new_x), int(new_y), int(new_w), int(new_h))


# ─── Image Launch Button ──────────────────────────────────────────────────────

class ImageLaunchButton(QPushButton):
    """A round, glowing image button.  Falls back to a painted '◈' crest
    if no image file is found next to the script."""

    def __init__(self, image_path: str = None, size: int = 130, parent=None):
        super().__init__(parent)
        self._size   = size
        self._pixmap = None
        self._hover  = False
        self._glow   = 0.0        # 0.0 → 1.0 driven by animation
        self.setFixedSize(size, size)
        self.setCursor(Qt.PointingHandCursor)
        self.setFlat(True)
        self.setStyleSheet("background: transparent; border: none;")

        # Try to load the image (looks for digi_valet_logo.png / .jpg next to script)
        candidates = []
        if image_path:
            candidates.append(Path(image_path))
        script_dir = Path(__file__).resolve().parent
        for name in ("digi_valet_logo.png", "digi_valet_logo.jpg",
                     "logo.png", "logo.jpg", "avatar.png", "avatar.jpg"):
            candidates.append(script_dir / name)

        for p in candidates:
            if p.exists():
                px = QPixmap(str(p))
                if not px.isNull():
                    self._pixmap = px.scaled(
                        size, size,
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                    break

        # Glow pulse animation
        self._glow_anim = QPropertyAnimation(self, b"glowValue", self)
        self._glow_anim.setStartValue(0.0)
        self._glow_anim.setEndValue(1.0)
        self._glow_anim.setDuration(1400)
        self._glow_anim.setEasingCurve(QEasingCurve.InOutSine)
        self._glow_anim.setLoopCount(-1)
        self._glow_anim.finished.connect(self._flip_glow)
        self._glow_anim.start()

    # Qt property so QPropertyAnimation can drive it
    def _get_glow(self): return self._glow
    def _set_glow(self, v):
        self._glow = v
        self.update()
    from PySide6.QtCore import Property as _Prop
    glowValue = _Prop(float, _get_glow, _set_glow)

    def _flip_glow(self):
        s, e = self._glow_anim.startValue(), self._glow_anim.endValue()
        self._glow_anim.setStartValue(e)
        self._glow_anim.setEndValue(s)
        self._glow_anim.start()

    def stop_animation(self):
        self._glow_anim.stop()

    def enterEvent(self, e):
        self._hover = True
        self.update()
        super().enterEvent(e)

    def leaveEvent(self, e):
        self._hover = False
        self.update()
        super().leaveEvent(e)

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        s = self._size
        cx, cy, r = s // 2, s // 2, s // 2 - 4

        # ── outer glow ring ──────────────────────────────────────────────
        glow_alpha = int(60 + self._glow * 120) + (40 if self._hover else 0)
        glow_r = r + 8 + int(self._glow * 6)
        grad = QRadialGradient(cx, cy, glow_r)
        grad.setColorAt(0.0, QColor(200, 169, 110, min(glow_alpha, 255)))
        grad.setColorAt(0.6, QColor(200, 169, 110, glow_alpha // 3))
        grad.setColorAt(1.0, QColor(200, 169, 110, 0))
        p.setBrush(QBrush(grad))
        p.setPen(Qt.NoPen)
        p.drawEllipse(cx - glow_r, cy - glow_r, glow_r * 2, glow_r * 2)

        # ── circular clip for image / fallback crest ─────────────────────
        path = QPainterPath()
        path.addEllipse(QRectF(cx - r, cy - r, r * 2, r * 2))
        p.setClipPath(path)

        if self._pixmap:
            # Draw image centred inside circle
            px_rect = self._pixmap.rect()
            draw_x = cx - px_rect.width() // 2
            draw_y = cy - px_rect.height() // 2
            p.drawPixmap(draw_x, draw_y, self._pixmap)
        else:
            # Fallback: dark gold background + '◈' crest
            bg = QLinearGradient(0, 0, 0, s)
            bg.setColorAt(0, QColor("#1e1a14"))
            bg.setColorAt(1, QColor("#0e0c0a"))
            p.fillRect(0, 0, s, s, QBrush(bg))
            p.setClipping(False)
            p.setPen(QColor("#c8a96e"))
            p.setFont(QFont("Georgia", int(s * 0.38), QFont.Bold))
            p.drawText(QRect(0, 0, s, s), Qt.AlignCenter, "◈")
            p.setClipPath(path)

        p.setClipping(False)

        # ── gold border ring ─────────────────────────────────────────────
        border_alpha = 180 + int(self._glow * 75)
        border_w = 2.5 + self._glow * 1.0
        p.setPen(QPen(QColor(200, 169, 110, min(border_alpha, 255)), border_w))
        p.setBrush(Qt.NoBrush)
        p.drawEllipse(QRectF(cx - r + 1, cy - r + 1, (r - 1) * 2, (r - 1) * 2))


# ─── Curtain Screen ───────────────────────────────────────────────────────────

class CurtainScreen(QWidget):
    """Full-window curtain overlay.

    Two drapes (left / right) start fully closed (meeting in the centre).
    Call open_curtain() to animate them sliding apart.
    Emits  opened  when the animation is complete.
    """

    opened = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        self.setCursor(Qt.PointingHandCursor)
        self._progress = 0.0    # 0.0 = fully closed, 1.0 = fully open
        self._animating = False
        self._dismissed = False
        self._anim = None

    # Qt property so QPropertyAnimation can drive _progress
    def _get_progress(self): return self._progress
    def _set_progress(self, v):
        self._progress = v
        self.update()
    from PySide6.QtCore import Property as _Prop2
    curtainProgress = _Prop2(float, _get_progress, _set_progress)

    def open_curtain(self):
        if self._animating:
            return
        self._animating = True
        self._dismissed = True
        self.setCursor(Qt.ArrowCursor)
        self._anim = QPropertyAnimation(self, b"curtainProgress", self)
        self._anim.setStartValue(0.0)
        self._anim.setEndValue(1.0)
        self._anim.setDuration(900)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)
        self._anim.finished.connect(self.opened.emit)
        self._anim.start()

    def mousePressEvent(self, event):
        if not self._dismissed:
            self.open_curtain()
        super().mousePressEvent(event)

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        half = w // 2
        # How far each drape has slid outward (0 px → fully open = half px)
        slide = int(self._progress * half)

        drape_color_top    = QColor("#1a1410")
        drape_color_bottom = QColor("#0a0806")
        drape_edge         = QColor("#c8a96e")
        shadow_color       = QColor(0, 0, 0, 160)

        for side in ("left", "right"):
            if side == "left":
                x0 = -slide
                drape_rect = QRect(x0, 0, half, h)
            else:
                x0 = half + slide
                drape_rect = QRect(x0, 0, half, h)

            # Fabric gradient
            grad = QLinearGradient(drape_rect.left(), 0, drape_rect.right(), 0)
            if side == "left":
                grad.setColorAt(0.0,  drape_color_top)
                grad.setColorAt(0.85, drape_color_bottom)
                grad.setColorAt(1.0,  QColor("#2a2018"))
            else:
                grad.setColorAt(0.0,  QColor("#2a2018"))
                grad.setColorAt(0.15, drape_color_bottom)
                grad.setColorAt(1.0,  drape_color_top)
            p.fillRect(drape_rect, QBrush(grad))

            # Vertical fabric folds (subtle lighter stripes)
            fold_alpha = int(18 * (1 - self._progress * 0.5))
            fold_pen = QPen(QColor(255, 240, 200, fold_alpha), 1)
            p.setPen(fold_pen)
            num_folds = 6
            for i in range(1, num_folds + 1):
                fx = drape_rect.left() + int(drape_rect.width() * i / (num_folds + 1))
                p.drawLine(fx, 0, fx, h)

            # Inner-edge gold trim line
            if side == "left":
                edge_x = x0 + half - 1
            else:
                edge_x = x0
            p.setPen(QPen(drape_edge, 2))
            p.drawLine(edge_x, 0, edge_x, h)

            # Drop-shadow cast by the edge onto the revealed area
            if self._progress > 0:
                shadow_w = int(40 * (1 - self._progress * 0.7))
                if shadow_w > 0:
                    if side == "left":
                        sg = QLinearGradient(edge_x, 0, edge_x + shadow_w, 0)
                    else:
                        sg = QLinearGradient(edge_x, 0, edge_x - shadow_w, 0)
                    sg.setColorAt(0.0, shadow_color)
                    sg.setColorAt(1.0, QColor(0, 0, 0, 0))
                    if side == "left":
                        p.fillRect(QRect(edge_x, 0, shadow_w, h), QBrush(sg))
                    else:
                        p.fillRect(QRect(edge_x - shadow_w, 0, shadow_w, h), QBrush(sg))

        # Gold centre-seam line (disappears as curtain opens)
        seam_alpha = int(200 * (1 - self._progress))
        if seam_alpha > 0:
            p.setPen(QPen(QColor(200, 169, 110, seam_alpha), 1))
            p.drawLine(half, 0, half, h)


# ─── Digi Valet Reveal Overlay ────────────────────────────────────────────────

class DigiValetRevealOverlay(QWidget):
    """Full-screen overlay that plays the DIGI VALET reveal animation.

    Sequence (all driven by a 16 ms timer at ~60 fps):
      Phase 0  (0 → 0.35)  — letters materialise one by one from centre outward,
                              each scaling up from 0 and fading in.
      Phase 1  (0.35→0.70) — golden shimmer sweep travels left→right across the text.
      Phase 2  (0.70→0.85) — tagline fades in below ("YOUR PERSONAL AI ASSISTANT").
      Phase 3  (0.85→1.00) — whole composition scales up slightly and fades out.

    Emits `finished` when the animation is done so the caller can remove the widget.
    """

    finished = Signal()

    _TITLE   = "DIGI VALET"
    _TAGLINE = "Y O U R   P E R S O N A L   A I   A S S I S T A N T"

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, True)   # clicks pass through
        self.setAttribute(Qt.WA_NoSystemBackground, True)

        self._t        = 0.0    # 0.0 → 1.0 overall progress
        self._done     = False
        self._shimmer  = 0.0    # 0.0→1.0 position of the shimmer band
        self._scale    = 1.0    # whole-comp scale during fade-out

        # Timer fires every ~16 ms (~60 fps)
        self._timer = QTimer(self)
        self._timer.setInterval(16)
        self._timer.timeout.connect(self._tick)
        self._timer.start()

    # ── internal helpers ──────────────────────────────────────────────────

    @staticmethod
    def _ease_out_cubic(x: float) -> float:
        return 1.0 - (1.0 - max(0.0, min(1.0, x))) ** 3

    @staticmethod
    def _ease_in_out_sine(x: float) -> float:
        import math
        return -(math.cos(math.pi * x) - 1) / 2

    def _tick(self):
        if self._done:
            return
        self._t += 0.008          # ~1.3 s total at 60 fps
        if self._t >= 1.0:
            self._t = 1.0
            self._done = True
            self._timer.stop()
            self.update()
            QTimer.singleShot(60, self.finished.emit)
            return
        self.update()

    # ── painting ──────────────────────────────────────────────────────────

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.setRenderHint(QPainter.TextAntialiasing)
        w, h = self.width(), self.height()
        t = self._t

        # ── background darkens in during first 15 % then holds ─────────────
        bg_alpha = int(min(t / 0.15, 1.0) * 210)
        p.fillRect(0, 0, w, h, QColor(8, 6, 4, bg_alpha))

        cx = w // 2
        title_y = h // 2 - 30          # vertical centre of the title
        tag_y   = title_y + 72         # tagline sits below

        title     = self._TITLE
        n_letters = len(title)

        # ── Phase 0: per-letter materialise (t = 0 → 0.38) ──────────────────
        phase0_end = 0.38
        # Each letter gets a staggered window
        letter_data = []   # (char, x_offset_from_centre, alpha, letter_scale)

        # Measure total width to centre
        font_title = QFont("Georgia", 54, QFont.Bold)
        p.setFont(font_title)
        fm = p.fontMetrics()
        total_w = sum(fm.horizontalAdvance(c) for c in title) + fm.horizontalAdvance(" ") * title.count(" ")
        # Accumulate x positions
        x_positions = []
        cur_x = cx - total_w // 2
        for ch in title:
            x_positions.append(cur_x)
            cur_x += fm.horizontalAdvance(ch)

        for i, ch in enumerate(title):
            # Letters in the middle reveal first, edges reveal later
            mid = (n_letters - 1) / 2.0
            dist_from_mid = abs(i - mid) / mid if mid > 0 else 0
            # stagger: middle letters start at t=0, edge letters start at t=0.15
            letter_start = 0.0 + dist_from_mid * 0.15
            letter_end   = letter_start + (phase0_end - 0.15)
            local_t = (t - letter_start) / max(letter_end - letter_start, 0.001)
            local_t = max(0.0, min(1.0, local_t))
            eased   = self._ease_out_cubic(local_t)
            alpha   = int(eased * 255)
            lscale  = 0.4 + eased * 0.6   # scale from 40 % → 100 %
            letter_data.append((ch, x_positions[i], alpha, lscale))

        # ── Phase 1: shimmer sweep (t = 0.35 → 0.72) ────────────────────────
        shimmer_local = (t - 0.35) / 0.37
        shimmer_local = max(0.0, min(1.0, shimmer_local))
        shimmer_pos   = self._ease_in_out_sine(shimmer_local)  # 0→1 across width

        # ── Phase 3: scale-up + fade-out (t = 0.82 → 1.0) ──────────────────
        fadeout_local = (t - 0.82) / 0.18
        fadeout_local = max(0.0, min(1.0, fadeout_local))
        fadeout_alpha = int((1.0 - fadeout_local) * 255)  # 255→0
        comp_scale    = 1.0 + fadeout_local * 0.12         # 1.0→1.12

        # Apply whole-composition transform around centre
        p.save()
        p.translate(cx, h // 2)
        p.scale(comp_scale, comp_scale)
        p.translate(-cx, -h // 2)

        # ── Draw each letter ──────────────────────────────────────────────────
        p.setFont(font_title)
        fm = p.fontMetrics()
        asc = fm.ascent()

        for ch, lx, alpha, lscale in letter_data:
            if alpha == 0:
                continue
            final_alpha = min(alpha, fadeout_alpha)

            # Save/restore to scale individual letter around its centre
            lw = fm.horizontalAdvance(ch)
            lcx = lx + lw // 2

            p.save()
            p.translate(lcx, title_y + asc // 2)
            p.scale(lscale, lscale)
            p.translate(-lcx, -(title_y + asc // 2))

            # Base gold colour
            base_col = QColor(200, 169, 110, final_alpha)
            p.setPen(base_col)
            p.drawText(lx, title_y + asc, ch)

            # Shimmer highlight: bright white-gold band
            if shimmer_local > 0 and lscale > 0.95:
                band_x   = int(shimmer_pos * (total_w + 200)) - 100 + (cx - total_w // 2)
                dist     = abs(lcx - band_x)
                band_w   = 90
                if dist < band_w:
                    shine   = 1.0 - (dist / band_w)
                    shine   = shine ** 2
                    shine_a = int(shine * min(230, fadeout_alpha))
                    p.setPen(QColor(255, 248, 220, shine_a))
                    p.drawText(lx, title_y + asc, ch)

            p.restore()

        # ── Tagline fade-in (t = 0.68 → 0.88) ───────────────────────────────
        tag_local = (t - 0.68) / 0.20
        tag_local = max(0.0, min(1.0, tag_local))
        tag_alpha = int(self._ease_out_cubic(tag_local) * min(180, fadeout_alpha + (255 - fadeout_alpha) * (1 - fadeout_local)))
        if tag_alpha > 0:
            font_tag = QFont("Georgia", 11)
            font_tag.setLetterSpacing(QFont.AbsoluteSpacing, 2)
            p.setFont(font_tag)
            p.setPen(QColor(138, 122, 106, tag_alpha))
            fm_tag = p.fontMetrics()
            tag_w  = fm_tag.horizontalAdvance(self._TAGLINE)
            p.drawText(cx - tag_w // 2, tag_y + fm_tag.ascent(), self._TAGLINE)

        # ── Decorative rule lines either side of DIGI VALET ─────────────────
        rule_alpha = int(min(t / 0.40, 1.0) * min(80, fadeout_alpha))
        if rule_alpha > 0:
            rule_y   = title_y + asc // 2
            rule_gap = 24
            rule_len = int(self._ease_out_cubic(min(t / 0.45, 1.0)) * (w // 2 - total_w // 2 - rule_gap - 20))
            if rule_len > 0:
                p.setPen(QPen(QColor(200, 169, 110, rule_alpha), 1))
                left_x1  = cx - total_w // 2 - rule_gap - rule_len
                left_x2  = cx - total_w // 2 - rule_gap
                right_x1 = cx + total_w // 2 + rule_gap
                right_x2 = cx + total_w // 2 + rule_gap + rule_len
                p.drawLine(left_x1,  rule_y, left_x2,  rule_y)
                p.drawLine(right_x1, rule_y, right_x2, rule_y)

        p.restore()  # end comp-scale transform


# ─── Splash / Welcome Screen ───────────────────────────────────────────────────

class SplashScreen(QWidget):
    """Entry screen with a glowing image button.
    Clicking the button flies to the CurtainScreen, which then opens
    on any click to reveal the main app."""

    # Emitted when the user has clicked through both screens
    activated = Signal(QPoint)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("splashScreen")
        self.setAttribute(Qt.WA_StyledBackground, True)
        self._dismissed = False

        # ── Stack: [0] entry page, [1] curtain page ───────────────────
        self._stack = QStackedWidget(self)
        self._stack.setStyleSheet("background: transparent;")

        # ── Page 0: the image-button entry screen ─────────────────────
        entry_page = QWidget()
        entry_page.setObjectName("splashScreen")
        entry_page.setAttribute(Qt.WA_StyledBackground, True)
        entry_layout = QVBoxLayout(entry_page)
        entry_layout.setAlignment(Qt.AlignCenter)
        entry_layout.setSpacing(0)

        logo = QLabel("DIGI VALET")
        logo.setObjectName("splashLogo")
        logo.setFont(QFont("Georgia", 44, QFont.Bold))
        logo.setAlignment(Qt.AlignCenter)
        entry_layout.addWidget(logo)

        entry_layout.addSpacing(6)

        tagline = QLabel("Y O U R   P E R S O N A L   A I   A S S I S T A N T")
        tagline.setObjectName("splashTagline")
        tagline.setFont(QFont("Georgia", 11))
        tagline.setAlignment(Qt.AlignCenter)
        entry_layout.addWidget(tagline)

        entry_layout.addSpacing(48)

        # ── Glowing image / crest button ──────────────────────────────
        self._launch_btn = ImageLaunchButton(size=130)
        self._launch_btn.clicked.connect(self._on_launch_clicked)
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setAlignment(Qt.AlignCenter)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.addWidget(self._launch_btn)
        entry_layout.addWidget(btn_container)

        entry_layout.addSpacing(18)

        self.hint = QLabel("✦  Tap the crest to enter  ✦")
        self.hint.setObjectName("splashHint")
        self.hint.setFont(QFont("Georgia", 10))
        self.hint.setAlignment(Qt.AlignCenter)
        entry_layout.addWidget(self.hint)

        entry_page.setStyleSheet(
            "#splashScreen { background: qlineargradient(x1:0, y1:0, x2:0, y2:1,"
            " stop:0 #0e0c0a, stop:1 #1a1612); }"
            "#splashLogo { color: #c8a96e; letter-spacing: 10px; }"
            "#splashTagline { color: #8a7a6a; letter-spacing: 2px; }"
            "#splashHint { color: #5a4f44; }"
        )

        # Hint pulse
        self._hint_effect = QGraphicsOpacityEffect(self.hint)
        self.hint.setGraphicsEffect(self._hint_effect)
        self._pulse_anim = QPropertyAnimation(self._hint_effect, b"opacity", self)
        self._pulse_anim.setStartValue(1.0)
        self._pulse_anim.setEndValue(0.30)
        self._pulse_anim.setDuration(1300)
        self._pulse_anim.setEasingCurve(QEasingCurve.InOutSine)
        self._pulse_anim.finished.connect(self._flip_pulse)
        self._pulse_anim.start()

        self._stack.addWidget(entry_page)   # index 0

        # ── Page 1: the dark curtain screen ───────────────────────────
        self._curtain = CurtainScreen()
        self._curtain.opened.connect(self._on_curtain_opened)
        self._stack.addWidget(self._curtain)  # index 1

        # Fill the whole SplashScreen widget
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(self._stack)

    # ── Helpers ────────────────────────────────────────────────────────

    def _flip_pulse(self):
        s, e = self._pulse_anim.startValue(), self._pulse_anim.endValue()
        self._pulse_anim.setStartValue(e)
        self._pulse_anim.setEndValue(s)
        self._pulse_anim.start()

    def _on_launch_clicked(self):
        """Button pressed → stop animations, zoom-out entry page, switch to curtain."""
        if self._dismissed:
            return
        self._dismissed = True
        self._pulse_anim.stop()
        self._launch_btn.stop_animation()

        # Brief zoom-out on the entry page then swap
        entry_page = self._stack.widget(0)
        effect = QGraphicsOpacityEffect(entry_page)
        entry_page.setGraphicsEffect(effect)

        fade = QPropertyAnimation(effect, b"opacity", self)
        fade.setStartValue(1.0)
        fade.setEndValue(0.0)
        fade.setDuration(260)
        fade.setEasingCurve(QEasingCurve.OutCubic)
        fade.finished.connect(self._show_curtain)
        self._fade_ref = fade  # prevent GC
        fade.start()

    def _show_curtain(self):
        self._stack.setCurrentIndex(1)   # switch to curtain page

    def _on_curtain_opened(self):
        """Curtain has fully opened — tell the main window to reveal itself."""
        centre = QPoint(self.width() // 2, self.height() // 2)
        self.activated.emit(centre)


# ─── Task Manager Dialog ──────────────────────────────────────────────────────

class TaskManagerWidget(QWidget):
    """Lightweight sidebar task panel."""
    tasks_changed = Signal()

    def __init__(self, tasks_file: Path, parent=None):
        super().__init__(parent)
        self.tasks_file = tasks_file
        self.tasks = self._load()
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        hdr = QLabel("MY TASKS")
        hdr.setObjectName("sectionLabel")
        hdr.setFont(QFont("Georgia", 8))
        layout.addWidget(hdr)

        self.list_widget = QListWidget()
        self.list_widget.setObjectName("historyList")
        self.list_widget.setFont(QFont("Georgia", 9))
        self.list_widget.setFixedHeight(130)
        self.list_widget.itemDoubleClicked.connect(self._toggle_task)
        layout.addWidget(self.list_widget)

        row = QHBoxLayout()
        self.task_input = QLineEdit()
        self.task_input.setObjectName("searchBox")
        self.task_input.setPlaceholderText("Add task…")
        self.task_input.setFont(QFont("Georgia", 9))
        self.task_input.returnPressed.connect(self._add_task)
        row.addWidget(self.task_input)

        add_btn = QPushButton("+")
        add_btn.setObjectName("newChatBtn")
        add_btn.setFixedWidth(28)
        add_btn.setFont(QFont("Georgia", 12, QFont.Bold))
        add_btn.clicked.connect(self._add_task)
        row.addWidget(add_btn)
        layout.addLayout(row)

        self._refresh_list()

    def _refresh_list(self):
        self.list_widget.clear()
        for t in self.tasks:
            prefix = "✓ " if t.get("done") else "○ "
            item = QListWidgetItem(prefix + t["text"])
            item.setForeground(QColor("#4a4038" if t.get("done") else "#c8a96e"))
            self.list_widget.addItem(item)

    def _add_task(self):
        text = self.task_input.text().strip()
        if text:
            self.tasks.append({"text": text, "done": False, "created": datetime.now().isoformat()})
            self.task_input.clear()
            self._refresh_list()
            self._save()
            self.tasks_changed.emit()

    def _toggle_task(self, item):
        row = self.list_widget.row(item)
        if 0 <= row < len(self.tasks):
            self.tasks[row]["done"] = not self.tasks[row]["done"]
            self._refresh_list()
            self._save()

    def _load(self):
        try:
            if self.tasks_file.exists():
                return json.loads(self.tasks_file.read_text())
        except Exception:
            pass
        return []

    def _save(self):
        try:
            self.tasks_file.write_text(json.dumps(self.tasks, indent=2))
        except Exception as e:
            print(f"Task save error: {e}")

    def get_task_summary(self) -> str:
        if not self.tasks:
            return "No tasks yet."
        lines = []
        for t in self.tasks:
            status = "✓" if t.get("done") else "○"
            lines.append(f"{status} {t['text']}")
        return "\n".join(lines)


# ─── App Icon (drawn in code — no external image needed) ──────────────────────

def _make_icon_pixmap(size: int = 64, dark: bool = True) -> QPixmap:
    """Draw the Digi Valet diamond ◈ icon at the given size."""
    gold = QColor("#c8a96e")
    bg = QColor("#141210") if dark else QColor("#fff8ec")

    px = QPixmap(size, size)
    px.fill(QColor(0, 0, 0, 0))
    p = QPainter(px)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)

    # Background circle
    p.setBrush(bg)
    p.setPen(gold)
    p.drawEllipse(2, 2, size - 4, size - 4)

    # Diamond shape
    cx, cy = size // 2, size // 2
    r = size // 3
    poly = QPolygon([
        QPoint(cx,     cy - r),
        QPoint(cx + r, cy),
        QPoint(cx,     cy + r),
        QPoint(cx - r, cy),
    ])
    p.setBrush(gold)
    p.setPen(QColor(0, 0, 0, 0))
    p.drawPolygon(poly)

    # Inner dot
    p.setBrush(bg)
    s = max(1, size // 10)
    p.drawEllipse(cx - s, cy - s, s * 2, s * 2)
    p.end()
    return px


def make_app_icon(dark: bool = True) -> QIcon:
    icon = QIcon()
    for sz in (16, 24, 32, 48, 64, 128, 256):
        icon.addPixmap(_make_icon_pixmap(sz, dark))
    return icon


# ─── Custom Title Bar (frameless window) ───────────────────────────────────────

class TitleBar(QWidget):
    """Custom drag-able title bar used when the window runs frameless."""

    def __init__(self, window: "DigiValetWindow"):
        super().__init__(window)
        self._win = window
        self._dragging = False
        self._drag_offset = QPoint()
        self.setFixedHeight(40)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 0, 8, 0)
        layout.setSpacing(8)

        self.icon_lbl = QLabel("◈")
        self.icon_lbl.setFont(QFont("Georgia", 14))
        layout.addWidget(self.icon_lbl)

        self.title_lbl = QLabel("DIGI VALET")
        self.title_lbl.setFont(QFont("Georgia", 11, QFont.Bold))
        layout.addWidget(self.title_lbl)

        self.tagline_lbl = QLabel("PERSONAL ASSISTANT")
        self.tagline_lbl.setFont(QFont("Georgia", 8))
        layout.addWidget(self.tagline_lbl)

        layout.addStretch()

        self.min_btn = QPushButton("─")
        self.max_btn = QPushButton("▢")
        self.close_btn = QPushButton("✕")
        for btn, tip, slot in [
            (self.min_btn,   "Minimise",         window.showMinimized),
            (self.max_btn,   "Maximise/Restore", self._toggle_max),
            (self.close_btn, "Close to tray",    window.close),
        ]:
            btn.setFixedSize(34, 28)
            btn.setFont(QFont("Georgia", 11))
            btn.setToolTip(tip)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(slot)
            layout.addWidget(btn)

        self.apply_theme(True)

    # ── Theming ──────────────────────────────────────────────────────────
    def apply_theme(self, dark: bool):
        if dark:
            bg, border, gold, text2 = "#141210", "#2a2420", "#c8a96e", "#8a7a6a"
        else:
            bg, border, gold, text2 = "#ede8dc", "#d8d0c0", "#8b6914", "#a09070"

        self.setStyleSheet(f"background:{bg}; border-bottom:1px solid {border};")
        self.icon_lbl.setStyleSheet(f"color:{gold}; background:transparent;")
        self.title_lbl.setStyleSheet(f"color:{gold}; letter-spacing:2px; background:transparent;")
        self.tagline_lbl.setStyleSheet(f"color:{text2}; letter-spacing:3px; background:transparent;")
        btn_style = (
            f"QPushButton {{ background:transparent; color:{text2}; border:none; border-radius:5px; }}"
            f"QPushButton:hover {{ background:{border}; color:{gold}; }}"
        )
        for btn in (self.min_btn, self.max_btn, self.close_btn):
            btn.setStyleSheet(btn_style)
        self.close_btn.setStyleSheet(
            btn_style + "QPushButton:hover { background:#e07070; color:#1a0a0a; }"
        )

    # ── Maximise / restore ──────────────────────────────────────────────────
    def _toggle_max(self):
        if self._win.isMaximized():
            self._win.showNormal()
            self.max_btn.setText("▢")
        else:
            self._win.showMaximized()
            self.max_btn.setText("❐")

    # ── Drag to move the frameless window ──────────────────────────────────
    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._drag_offset = e.globalPosition().toPoint() - self._win.frameGeometry().topLeft()

    def mouseMoveEvent(self, e):
        if self._dragging and e.buttons() == Qt.MouseButton.LeftButton:
            if self._win.isMaximized():
                self._win.showNormal()
                self.max_btn.setText("▢")
            self._win.move(e.globalPosition().toPoint() - self._drag_offset)

    def mouseReleaseEvent(self, e):
        self._dragging = False

    def mouseDoubleClickEvent(self, e):
        self._toggle_max()


# ─── Stylesheets ──────────────────────────────────────────────────────────────

DARK_STYLESHEET = """
QWidget#windowShell { background-color: #0e0c0a; border: 1px solid #2a2420; }
QMainWindow, QWidget#centralWidget { background-color: #0e0c0a; }
QWidget#sidebar { background-color: #141210; border-right: 1px solid #2a2420; }
QLabel#logoLabel { color: #c8a96e; font-size: 22px; letter-spacing: 3px; }
QLabel#taglineLabel { color: #5c5046; font-size: 10px; letter-spacing: 4px; }
QLabel#sectionLabel { color: #4a4038; font-size: 9px; letter-spacing: 3px; }
QLabel#modelLabel { color: #8a7a6a; font-size: 10px; }
QComboBox#modelCombo, QComboBox#toneCombo, QComboBox#langCombo {
    background-color: #1e1a16; color: #c8a96e; border: 1px solid #2e2820;
    border-radius: 6px; padding: 6px 10px; font-size: 11px; min-height: 28px;
}
QComboBox#modelCombo::drop-down, QComboBox#toneCombo::drop-down, QComboBox#langCombo::drop-down { border: none; width: 20px; }
QComboBox QAbstractItemView { background-color: #1e1a16; color: #c8a96e; border: 1px solid #3a3028; selection-background-color: #2e2820; }
QPushButton#newChatBtn {
    background-color: #1e1a16; color: #c8a96e; border: 1px solid #3a3028;
    border-radius: 6px; padding: 8px 14px; font-size: 11px; text-align: left;
}
QPushButton#newChatBtn:hover { background-color: #2a2420; border-color: #c8a96e; }
QPushButton#clearBtn {
    background-color: transparent; color: #5c5046; border: 1px solid #2a2420;
    border-radius: 6px; padding: 6px 14px; font-size: 10px;
}
QPushButton#clearBtn:hover { color: #e07070; border-color: #e07070; }
QListWidget#historyList { background-color: transparent; border: none; color: #8a7a6a; }
QListWidget#historyList::item { background-color: transparent; padding: 8px 10px; border-radius: 6px; margin-bottom: 4px; }
QListWidget#historyList::item:hover { background-color: #1e1a16; color: #c8a96e; }
QListWidget#historyList::item:selected { background-color: #2a2420; color: #c8a96e; font-weight: bold; }
QScrollArea#chatScroll { background-color: transparent; border: none; }
QWidget#chatContainer { background-color: transparent; }
QFrame#assistantBubble { background-color: #181410; border: 1px solid #2a2420; border-left: 3px solid #c8a96e; border-radius: 10px; margin: 3px 60px 3px 0px; }
QFrame#userBubble { background-color: #111820; border: 1px solid #1e2a38; border-right: 3px solid #8ab4c8; border-radius: 10px; margin: 3px 0px 3px 60px; }
QFrame#typingIndicator { background-color: #181410; border: 1px solid #2a2420; border-left: 3px solid #c8a96e; border-radius: 10px; margin: 3px 60px 3px 0px; }
QLabel#msgText { color: #d4cfc8; line-height: 1.6; background: transparent; }
QLabel#timeLabel { color: #3a3028; }
QWidget#inputArea { background-color: #0e0c0a; }
QWidget#inputRow { background-color: #161210; border: 1px solid #2a2420; border-radius: 22px; }
QWidget#inputRow[focused="true"] { border-color: #c8a96e; background-color: #1a1612; }
QTextEdit#inputBox { background-color: transparent; color: #d4cfc8; border: none; padding: 9px 4px; font-size: 13px; selection-background-color: #3a3028; }
QPushButton#sendBtn { background-color: #c8a96e; color: #0e0c0a; border: none; border-radius: 20px; font-size: 16px; font-weight: bold; min-width: 40px; min-height: 40px; max-width: 40px; max-height: 40px; }
QPushButton#sendBtn:hover { background-color: #dbbf82; }
QPushButton#sendBtn:disabled { background-color: #2a2420; color: #4a4038; }
QPushButton#attachBtn { background-color: transparent; color: #c8a96e; border: none; border-radius: 20px; font-size: 16px; font-weight: bold; min-width: 40px; max-width: 40px; min-height: 40px; max-height: 40px; }
QPushButton#attachBtn:hover { background-color: #2a2420; }
QPushButton#attachBtn[hasFiles="true"] { background-color: #2a2015; color: #e8c97e; }
QPushButton#stopBtn { background-color: #3a1818; color: #e07070; border: none; border-radius: 20px; font-size: 12px; min-width: 40px; min-height: 40px; max-width: 40px; max-height: 40px; }
QPushButton#stopBtn:hover { background-color: #4a2020; }
QLabel#statusLabel { color: #4a4038; font-size: 10px; letter-spacing: 1px; }
QPushButton#copyBtn { background: transparent; color: #4a4038; border: none; border-radius: 4px; font-size: 12px; padding: 0px; }
QPushButton#copyBtn:hover { color: #c8a96e; background: #2a2420; }
QPushButton#fontBtn { background-color: #1e1a16; color: #8a7a6a; border: 1px solid #2a2420; border-radius: 5px; font-size: 13px; font-weight: bold; min-width: 26px; max-width: 26px; min-height: 26px; max-height: 26px; }
QPushButton#fontBtn:hover { background-color: #2a2420; color: #c8a96e; border-color: #c8a96e; }
QPushButton#exportBtn { background-color: transparent; color: #5c5046; border: 1px solid #2a2420; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#exportBtn:hover { color: #c8a96e; border-color: #c8a96e; }
QPushButton#themeBtn { background-color: #1e1a16; color: #8a7a6a; border: 1px solid #2a2420; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#themeBtn:hover { color: #c8a96e; border-color: #c8a96e; }
QPushButton#privacyBtn { background-color: transparent; color: #5c5046; border: 1px solid #2a2420; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#privacyBtn:hover { color: #8ab4c8; border-color: #8ab4c8; }
QLineEdit#searchBox { background-color: #1a1612; color: #c8a96e; border: 1px solid #2a2420; border-radius: 6px; padding: 5px 10px; font-size: 10px; selection-background-color: #3a3028; }
QLineEdit#searchBox:focus { border-color: #c8a96e; }
QScrollBar:vertical { background: transparent; width: 6px; margin: 0; }
QScrollBar::handle:vertical { background: #2a2420; border-radius: 3px; min-height: 20px; }
QScrollBar::handle:vertical:hover { background: #c8a96e; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: transparent; }
"""

LIGHT_STYLESHEET = """
QWidget#windowShell { background-color: #f5f0e8; border: 1px solid #d8d0c0; }
QMainWindow, QWidget#centralWidget { background-color: #f5f0e8; }
QWidget#sidebar { background-color: #ede8dc; border-right: 1px solid #d0c8b8; }
QLabel#logoLabel { color: #8b6914; font-size: 22px; letter-spacing: 3px; }
QLabel#taglineLabel { color: #a09070; font-size: 10px; letter-spacing: 4px; }
QLabel#sectionLabel { color: #b0a080; font-size: 9px; letter-spacing: 3px; }
QLabel#modelLabel { color: #705840; font-size: 10px; }
QComboBox#modelCombo, QComboBox#toneCombo, QComboBox#langCombo {
    background-color: #fff8ec; color: #8b6914; border: 1px solid #d0b870;
    border-radius: 6px; padding: 6px 10px; font-size: 11px; min-height: 28px;
}
QComboBox#modelCombo::drop-down, QComboBox#toneCombo::drop-down, QComboBox#langCombo::drop-down { border: none; width: 20px; }
QComboBox QAbstractItemView { background-color: #fff8ec; color: #8b6914; border: 1px solid #d0b870; selection-background-color: #f0e0b0; }
QPushButton#newChatBtn {
    background-color: #fff8ec; color: #8b6914; border: 1px solid #d0b870;
    border-radius: 6px; padding: 8px 14px; font-size: 11px; text-align: left;
}
QPushButton#newChatBtn:hover { background-color: #f0e8d0; border-color: #8b6914; }
QPushButton#clearBtn {
    background-color: transparent; color: #a09080; border: 1px solid #d0c8b8;
    border-radius: 6px; padding: 6px 14px; font-size: 10px;
}
QPushButton#clearBtn:hover { color: #c04040; border-color: #c04040; }
QListWidget#historyList { background-color: transparent; border: none; color: #706050; }
QListWidget#historyList::item { background-color: transparent; padding: 8px 10px; border-radius: 6px; margin-bottom: 4px; }
QListWidget#historyList::item:hover { background-color: #f0e8d0; color: #8b6914; }
QListWidget#historyList::item:selected { background-color: #e8d8b0; color: #6b4a00; font-weight: bold; }
QScrollArea#chatScroll { background-color: transparent; border: none; }
QWidget#chatContainer { background-color: transparent; }
QFrame#assistantBubble { background-color: #fff8ec; border: 1px solid #d0c8b0; border-left: 3px solid #c8a96e; border-radius: 10px; margin: 3px 60px 3px 0px; }
QFrame#userBubble { background-color: #eaf4fb; border: 1px solid #c0d8e8; border-right: 3px solid #6a9ab8; border-radius: 10px; margin: 3px 0px 3px 60px; }
QFrame#typingIndicator { background-color: #fff8ec; border: 1px solid #d0c8b0; border-left: 3px solid #c8a96e; border-radius: 10px; margin: 3px 60px 3px 0px; }
QLabel#msgText { color: #2a2010; line-height: 1.6; background: transparent; }
QLabel#timeLabel { color: #c0b090; }
QWidget#inputArea { background-color: #f5f0e8; }
QWidget#inputRow { background-color: #ffffff; border: 1px solid #d8d0c0; border-radius: 22px; }
QWidget#inputRow[focused="true"] { border-color: #c8a96e; background-color: #fffdf5; }
QTextEdit#inputBox { background-color: transparent; color: #2a2010; border: none; padding: 9px 4px; font-size: 13px; selection-background-color: #f0d890; }
QPushButton#sendBtn { background-color: #c8a96e; color: #fff8ec; border: none; border-radius: 20px; font-size: 16px; font-weight: bold; min-width: 40px; min-height: 40px; max-width: 40px; max-height: 40px; }
QPushButton#sendBtn:hover { background-color: #dbbf82; }
QPushButton#sendBtn:disabled { background-color: #d8d0c0; color: #a09080; }
QPushButton#attachBtn { background-color: transparent; color: #8b6914; border: none; border-radius: 20px; font-size: 16px; font-weight: bold; min-width: 40px; max-width: 40px; min-height: 40px; max-height: 40px; }
QPushButton#attachBtn:hover { background-color: #f0e8d0; }
QPushButton#attachBtn[hasFiles="true"] { background-color: #fff0cc; color: #7a5900; }
QPushButton#stopBtn { background-color: #fce8e8; color: #c04040; border: none; border-radius: 20px; font-size: 12px; min-width: 40px; min-height: 40px; max-width: 40px; max-height: 40px; }
QPushButton#stopBtn:hover { background-color: #f8d8d8; }
QLabel#statusLabel { color: #b0a080; font-size: 10px; letter-spacing: 1px; }
QPushButton#copyBtn { background: transparent; color: #b0a080; border: none; border-radius: 4px; font-size: 12px; padding: 0px; }
QPushButton#copyBtn:hover { color: #8b6914; background: #f0e0b0; }
QPushButton#fontBtn { background-color: #fff8ec; color: #706050; border: 1px solid #d0c8b0; border-radius: 5px; font-size: 13px; font-weight: bold; min-width: 26px; max-width: 26px; min-height: 26px; max-height: 26px; }
QPushButton#fontBtn:hover { background-color: #f0e8d0; color: #8b6914; border-color: #c8a96e; }
QPushButton#exportBtn { background-color: transparent; color: #a09080; border: 1px solid #d0c8b0; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#exportBtn:hover { color: #8b6914; border-color: #c8a96e; }
QPushButton#themeBtn { background-color: #fff8ec; color: #706050; border: 1px solid #d0c8b0; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#themeBtn:hover { color: #8b6914; border-color: #c8a96e; }
QPushButton#privacyBtn { background-color: transparent; color: #a09080; border: 1px solid #d0c8b0; border-radius: 6px; padding: 5px 10px; font-size: 10px; }
QPushButton#privacyBtn:hover { color: #4a7090; border-color: #6a9ab8; }
QLineEdit#searchBox { background-color: #fff8ec; color: #8b6914; border: 1px solid #d0c8b0; border-radius: 6px; padding: 5px 10px; font-size: 10px; selection-background-color: #f0d890; }
QLineEdit#searchBox:focus { border-color: #c8a96e; }
QScrollBar:vertical { background: transparent; width: 6px; margin: 0; }
QScrollBar::handle:vertical { background: #d0c8b0; border-radius: 3px; min-height: 20px; }
QScrollBar::handle:vertical:hover { background: #c8a96e; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: transparent; }
"""


# ─── App Container (manages splash + main page stacking) ──────────────────────

class _AppContainer(QWidget):
    """Central widget that hosts both the splash screen and the main app page
    as freely-positioned children, so we can animate their geometry directly."""

    resized = Signal()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.resized.emit()


# ─── Main Window ──────────────────────────────────────────────────────────────

class DigiValetWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Digi Valet")
        self.resize(1020, 740)
        self.setMinimumSize(720, 500)

        # Frameless window with a custom title bar + tray icon
        self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        self.setWindowIcon(make_app_icon(True))
        self._tray = None
        self._force_quit = False

        self.chat_history_file = Path.home() / ".digi_valet_history_v2.json"
        self.tasks_file = Path.home() / ".digi_valet_tasks.json"
        self.prefs_file = Path.home() / ".digi_valet_prefs.json"

        # Load preferences
        self._prefs = self._load_prefs()
        global _DARK_MODE
        _DARK_MODE = self._prefs.get("dark_mode", True)

        self.all_chats = self._load_all_chats()
        self.current_chat_id = None
        self.messages = []
        self.privacy_mode = False

        self.current_worker = None
        self.current_thread = None
        self.current_bubble = None
        self.typing_indicator = None
        self.ollama_host = "http://localhost:11434"
        self._font_size_delta = 0

        # ── File attachment state ───────────────────────────────────────
        self._attached_files: list = []   # list of file paths pending send
        self._file_chips: dict = {}       # path → FileChip widget

        # ── Scroll behaviour ────────────────────────────────────────────
        # True = keep following the bottom during generation.
        # Flips to False the moment the user drags the scrollbar up.
        # Resets to True when the user sends a new message.
        self._auto_scroll: bool = True

        # ── Org Tree (offline org-chart lookup) ────────────────────────
        self.org_tree = OrgTree()
        self._load_org_tree()

        # ── Persistent Knowledge Base ───────────────────────────────
        self.kb = KnowledgeBase() if KnowledgeBase else None
        if self.kb and not self.kb.is_empty():
            files = self.kb.list_files()
            print(f"📚 Knowledge base loaded: {len(files)} file(s) — "
                  + ", ".join(f['filename'] for f in files))

        self._transition_active = False
        self._reveal = None
        self._build_ui()
        self._apply_theme()
        self._setup_tray()
        self._setup_shortcuts()
        self._load_models()
        self._initialize_default_chat()

    # ── Preferences ────────────────────────────────────────────────────────

    def _load_prefs(self) -> dict:
        try:
            if self.prefs_file.exists():
                return json.loads(self.prefs_file.read_text())
        except Exception:
            pass
        return {"dark_mode": True, "tone": "balanced", "language": "English"}

    def _save_prefs(self):
        try:
            self.prefs_file.write_text(json.dumps(self._prefs, indent=2))
        except Exception as e:
            print(f"Pref save error: {e}")

    # ── Org Tree ───────────────────────────────────────────────────────────

    def _load_org_tree(self):
        """Locate and load OrgTree.csv from several common locations."""
        import sys as _sys
        candidates = [
            Path(__file__).resolve().parent / "OrgTree.csv",
            Path(_sys.argv[0]).resolve().parent / "OrgTree.csv",
            Path.cwd() / "OrgTree.csv",
            Path.home() / "OrgTree.csv",
            Path.home() / ".digi_valet_orgtree.csv",
        ]
        for path in candidates:
            if path.exists():
                msg = self.org_tree.load_from_csv(str(path))
                print(msg)
                return
        print("ℹ️ OrgTree.csv not found — /digivalet will report 'not loaded' until you add one.")

    # ── UI Construction ────────────────────────────────────────────────────

    def _build_ui(self):
        # Outer shell: custom title bar on top, app container below
        shell = QWidget()
        shell.setObjectName("windowShell")
        shell_layout = QVBoxLayout(shell)
        shell_layout.setContentsMargins(0, 0, 0, 0)
        shell_layout.setSpacing(0)

        self.title_bar = TitleBar(self)
        shell_layout.addWidget(self.title_bar)

        self._app_container = _AppContainer()
        self._app_container.setObjectName("appContainer")
        shell_layout.addWidget(self._app_container)
        self._app_container.resized.connect(self._sync_app_geometry)

        self.setCentralWidget(shell)

        # Main application page (sidebar + chat) — starts hidden behind the splash
        self.main_widget = QWidget(self._app_container)
        self.main_widget.setObjectName("centralWidget")
        root = QHBoxLayout(self.main_widget)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_sidebar())
        root.addWidget(self._build_chat_area())
        self.main_widget.hide()

        # Welcome / splash screen — shown on top until the user clicks
        self.splash = SplashScreen(self._app_container)
        self.splash.activated.connect(self._reveal_main_app)
        self.splash.raise_()

        # Resize grip for the frameless window (bottom-right corner)
        self.size_grip = QSizeGrip(self)
        self.size_grip.setStyleSheet("background: transparent;")
        self.size_grip.setFixedSize(16, 16)
        self.size_grip.raise_()

        self._sync_app_geometry()

    def _sync_app_geometry(self):
        """Keep splash + main page filling the container (skipped mid-animation)."""
        if self._transition_active:
            return
        rect = self._app_container.rect()
        self.main_widget.setGeometry(rect)
        self.splash.setGeometry(rect)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        # Keep the resize grip pinned to the bottom-right corner
        if hasattr(self, "size_grip"):
            g = self.size_grip
            g.move(self.width() - g.width(), self.height() - g.height())
            g.setVisible(not self.isMaximized() and not self.isFullScreen())
            g.raise_()

    # ── Splash → Main transition ──────────────────────────────────────────

    def _reveal_main_app(self, click_pos: QPoint):
        """Curtain opened — play DIGI VALET reveal, then fade the main app in."""
        self._transition_active = True
        rect = self._app_container.rect()

        # Prepare main widget (invisible until reveal finishes)
        self.main_widget.setGeometry(rect)
        self.main_widget.hide()

        # Fade the splash (curtain) out quickly
        splash_effect = QGraphicsOpacityEffect(self.splash)
        self.splash.setGraphicsEffect(splash_effect)
        splash_fade = QPropertyAnimation(splash_effect, b"opacity", self)
        splash_fade.setStartValue(1.0)
        splash_fade.setEndValue(0.0)
        splash_fade.setDuration(350)
        splash_fade.setEasingCurve(QEasingCurve.OutCubic)
        splash_fade.finished.connect(self._start_reveal_overlay)
        self._splash_fade_ref = splash_fade
        splash_fade.start()

    def _start_reveal_overlay(self):
        """Splash gone — play the full-screen DIGI VALET reveal animation."""
        self.splash.hide()
        self.splash.setGraphicsEffect(None)

        # Show main app behind the overlay so it's ready
        rect = self._app_container.rect()
        self.main_widget.setGeometry(rect)
        main_effect = QGraphicsOpacityEffect(self.main_widget)
        main_effect.setOpacity(0.0)
        self.main_widget.setGraphicsEffect(main_effect)
        self.main_widget.show()
        self._main_effect_ref = main_effect

        # Create and show the reveal overlay on top
        self._reveal = DigiValetRevealOverlay(self._app_container)
        self._reveal.setGeometry(rect)
        self._reveal.show()
        self._reveal.raise_()
        self._reveal.finished.connect(self._on_reveal_finished)

    def _on_reveal_finished(self):
        """Reveal animation done — fade the main app to full opacity."""
        self._reveal.hide()
        self._reveal.deleteLater()
        self._reveal = None

        main_effect = self._main_effect_ref
        fade_in = QPropertyAnimation(main_effect, b"opacity", self)
        fade_in.setStartValue(0.0)
        fade_in.setEndValue(1.0)
        fade_in.setDuration(480)
        fade_in.setEasingCurve(QEasingCurve.OutCubic)
        fade_in.finished.connect(self._finish_transition)
        self._main_fadein_ref = fade_in
        fade_in.start()

    def _finish_transition(self):
        self.main_widget.setGraphicsEffect(None)
        self._transition_active = False
        self._sync_app_geometry()
        self.input_box.setFocus()

    def _build_sidebar(self):
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(250)
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(20, 28, 20, 20)
        layout.setSpacing(0)

        # Logo
        logo = QLabel("DIGI VALET")
        logo.setObjectName("logoLabel")
        logo.setFont(QFont("Georgia", 14, QFont.Bold))
        logo.setAlignment(Qt.AlignLeft)
        layout.addWidget(logo)

        tagline = QLabel("PERSONAL ASSISTANT")
        tagline.setObjectName("taglineLabel")
        tagline.setFont(QFont("Georgia", 8))
        layout.addWidget(tagline)

        layout.addSpacing(20)

        # Model selector
        model_sec = QLabel("MODEL")
        model_sec.setObjectName("sectionLabel")
        model_sec.setFont(QFont("Georgia", 8))
        layout.addWidget(model_sec)
        layout.addSpacing(6)

        self.model_combo = QComboBox()
        self.model_combo.setObjectName("modelCombo")
        self.model_combo.addItem("Loading models…")
        self.model_combo.setFont(QFont("Georgia", 10))
        layout.addWidget(self.model_combo)

        layout.addSpacing(14)

        # ── Tone selector ──
        tone_sec = QLabel("TONE")
        tone_sec.setObjectName("sectionLabel")
        tone_sec.setFont(QFont("Georgia", 8))
        layout.addWidget(tone_sec)
        layout.addSpacing(6)

        self.tone_combo = QComboBox()
        self.tone_combo.setObjectName("toneCombo")
        self.tone_combo.addItems(["Formal", "Balanced", "Casual"])
        self.tone_combo.setFont(QFont("Georgia", 10))
        saved_tone = self._prefs.get("tone", "balanced").capitalize()
        idx = self.tone_combo.findText(saved_tone)
        if idx >= 0:
            self.tone_combo.setCurrentIndex(idx)
        self.tone_combo.currentTextChanged.connect(self._on_tone_changed)
        layout.addWidget(self.tone_combo)

        layout.addSpacing(14)

        # ── Language selector ──
        lang_sec = QLabel("LANGUAGE")
        lang_sec.setObjectName("sectionLabel")
        lang_sec.setFont(QFont("Georgia", 8))
        layout.addWidget(lang_sec)
        layout.addSpacing(6)

        self.lang_combo = QComboBox()
        self.lang_combo.setObjectName("langCombo")
        self.lang_combo.addItems(list(LANGUAGE_ADDONS.keys()))
        self.lang_combo.setFont(QFont("Georgia", 10))
        saved_lang = self._prefs.get("language", "English")
        idx2 = self.lang_combo.findText(saved_lang)
        if idx2 >= 0:
            self.lang_combo.setCurrentIndex(idx2)
        self.lang_combo.currentTextChanged.connect(self._on_language_changed)
        layout.addWidget(self.lang_combo)

        layout.addSpacing(16)

        # New Chat button
        new_chat_btn = QPushButton("＋  New Conversation")
        new_chat_btn.setObjectName("newChatBtn")
        new_chat_btn.setFont(QFont("Georgia", 10))
        new_chat_btn.clicked.connect(self._new_chat)
        layout.addWidget(new_chat_btn)

        layout.addSpacing(14)

        # ── Task Manager ──
        self.task_widget = TaskManagerWidget(self.tasks_file)
        layout.addWidget(self.task_widget)

        layout.addSpacing(14)

        # Chat History List
        history_sec = QLabel("PAST CONVERSATIONS")
        history_sec.setObjectName("sectionLabel")
        history_sec.setFont(QFont("Georgia", 8))
        layout.addWidget(history_sec)
        layout.addSpacing(6)

        self.search_box = QLineEdit()
        self.search_box.setObjectName("searchBox")
        self.search_box.setPlaceholderText("🔍  Search conversations…")
        self.search_box.setFont(QFont("Georgia", 9))
        self.search_box.textChanged.connect(self._filter_sidebar)
        layout.addWidget(self.search_box)
        layout.addSpacing(6)

        self.history_list_widget = QListWidget()
        self.history_list_widget.setObjectName("historyList")
        self.history_list_widget.setFont(QFont("Georgia", 10))
        self.history_list_widget.itemClicked.connect(self._on_history_item_clicked)
        self.history_list_widget.itemDoubleClicked.connect(self._on_history_item_double_clicked)
        layout.addWidget(self.history_list_widget)

        layout.addSpacing(12)

        # Theme + Privacy row
        ctrl_row = QHBoxLayout()
        self.theme_btn = QPushButton("☀ Light" if _DARK_MODE else "☾ Dark")
        self.theme_btn.setObjectName("themeBtn")
        self.theme_btn.setFont(QFont("Georgia", 9))
        self.theme_btn.clicked.connect(self._toggle_theme)
        ctrl_row.addWidget(self.theme_btn)

        self.privacy_btn = QPushButton("🔒 Private")
        self.privacy_btn.setObjectName("privacyBtn")
        self.privacy_btn.setFont(QFont("Georgia", 9))
        self.privacy_btn.setCheckable(True)
        self.privacy_btn.clicked.connect(self._toggle_privacy)
        ctrl_row.addWidget(self.privacy_btn)
        layout.addLayout(ctrl_row)

        layout.addSpacing(8)

        # Clear All + Status
        clear_btn = QPushButton("Clear All History")
        clear_btn.setObjectName("clearBtn")
        clear_btn.setFont(QFont("Georgia", 9))
        clear_btn.clicked.connect(self._clear_all_chats)
        layout.addWidget(clear_btn)

        layout.addSpacing(10)

        self.sidebar_status = QLabel("● Checking Ollama…")
        self.sidebar_status.setObjectName("statusLabel")
        self.sidebar_status.setFont(QFont("Georgia", 9))
        layout.addWidget(self.sidebar_status)

        return sidebar

    def _build_chat_area(self):
        chat_area = QWidget()
        chat_layout = QVBoxLayout(chat_area)
        chat_layout.setContentsMargins(0, 0, 0, 0)
        chat_layout.setSpacing(0)

        # Header bar
        header = QWidget()
        header.setFixedHeight(56)
        header.setStyleSheet("background:#0e0c0a; border-bottom: 1px solid #1e1a16;")
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(24, 0, 24, 0)
        h_layout.setSpacing(10)

        title = QLabel("Conversation")
        title.setFont(QFont("Georgia", 13))
        title.setStyleSheet("color: #8a7a6a;")
        h_layout.addWidget(title)
        h_layout.addStretch()

        font_minus = QPushButton("A−")
        font_minus.setObjectName("fontBtn")
        font_minus.setFont(QFont("Georgia", 9))
        font_minus.setToolTip("Decrease font size")
        font_minus.clicked.connect(lambda: self._change_font_size(-1))
        h_layout.addWidget(font_minus)

        font_plus = QPushButton("A+")
        font_plus.setObjectName("fontBtn")
        font_plus.setFont(QFont("Georgia", 9))
        font_plus.setToolTip("Increase font size")
        font_plus.clicked.connect(lambda: self._change_font_size(+1))
        h_layout.addWidget(font_plus)

        h_layout.addSpacing(8)

        export_btn = QPushButton("⬇ Export")
        export_btn.setObjectName("exportBtn")
        export_btn.setFont(QFont("Georgia", 9))
        export_btn.setToolTip("Export current conversation")
        export_btn.clicked.connect(self._export_chat)
        h_layout.addWidget(export_btn)

        h_layout.addSpacing(8)

        self.model_indicator = QLabel("—")
        self.model_indicator.setFont(QFont("Georgia", 9))
        self.model_indicator.setStyleSheet("color: #3a3028; letter-spacing: 1px;")
        h_layout.addWidget(self.model_indicator)

        chat_layout.addWidget(header)

        # Quick commands bar
        cmd_bar = QWidget()
        cmd_bar.setFixedHeight(38)
        cmd_bar.setStyleSheet("background: #141210; border-bottom: 1px solid #1e1a16;" if _DARK_MODE else "background: #ede8dc; border-bottom: 1px solid #d8d0c0;")
        cmd_layout = QHBoxLayout(cmd_bar)
        cmd_layout.setContentsMargins(16, 4, 16, 4)
        cmd_layout.setSpacing(6)
        cmds = ["/help", "/tasks", "/plan", "/wellness", "/meal", "/summarize", "/focus", "/digivalet"]
        for cmd in cmds:
            btn = QPushButton(cmd)
            btn.setFont(QFont("Georgia", 8))
            btn.setStyleSheet(
                "QPushButton { background:#1e1a16; color:#8a7a6a; border:1px solid #2a2420; border-radius:4px; padding:2px 8px; }"
                "QPushButton:hover { color:#c8a96e; border-color:#c8a96e; }"
            )
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda checked, c=cmd: self._run_quick_command(c))
            cmd_layout.addWidget(btn)
        cmd_layout.addStretch()
        self.cmd_bar = cmd_bar
        chat_layout.addWidget(cmd_bar)

        # Scrollable chat
        self.scroll_area = QScrollArea()
        self.scroll_area.setObjectName("chatScroll")
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.chat_container = QWidget()
        self.chat_container.setObjectName("chatContainer")
        self.chat_layout = QVBoxLayout(self.chat_container)
        self.chat_layout.setContentsMargins(24, 20, 24, 20)
        self.chat_layout.setSpacing(10)
        self.chat_layout.addStretch()

        self.scroll_area.setWidget(self.chat_container)
        chat_layout.addWidget(self.scroll_area, 1)

        # Detect when user manually scrolls up — pause auto-scroll
        self.scroll_area.verticalScrollBar().valueChanged.connect(self._on_scroll_value_changed)

        chat_layout.addWidget(self._build_input_area(), 0)
        return chat_area

    def _build_input_area(self):
        # Outer wrapper — height is driven entirely by its children
        # (chips strip + pill), and locked with a vertical size policy of
        # Fixed/Minimum so leftover layout space never stretches it.
        outer = QWidget()
        outer.setObjectName("inputArea")
        outer.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)
        outer_layout = QVBoxLayout(outer)
        outer_layout.setContentsMargins(16, 8, 16, 14)
        outer_layout.setSpacing(8)

        # ── File chips strip (hidden when no files attached) ─────────────
        self._chips_bar = QWidget()
        self._chips_bar.setObjectName("chipsBar")
        self._chips_bar.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self._chips_bar.setFixedHeight(FileChip.CARD_HEIGHT)
        self._chips_layout = QHBoxLayout(self._chips_bar)
        self._chips_layout.setContentsMargins(0, 0, 0, 0)
        self._chips_layout.setSpacing(8)
        self._chips_layout.addStretch()
        self._chips_bar.hide()
        outer_layout.addWidget(self._chips_bar)

        # ── Input row (compact, Gemini-style pill bar) ─────────────────────
        input_row = QWidget()
        input_row.setObjectName("inputRow")
        input_row.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        layout = QHBoxLayout(input_row)
        layout.setContentsMargins(14, 6, 8, 6)
        layout.setSpacing(6)
        layout.setAlignment(Qt.AlignVCenter)

        # ── Attach button (Gemini-style) ─────────────────────────────────
        self.attach_btn = QPushButton("+")
        self.attach_btn.setObjectName("attachBtn")
        self.attach_btn.setFont(QFont("Georgia", 15, QFont.Bold))
        self.attach_btn.setFixedSize(40, 40)
        self.attach_btn.setToolTip("Attach files (any type, up to 100 files / 10 GB)")
        self.attach_btn.setCursor(Qt.PointingHandCursor)
        self.attach_btn.clicked.connect(self._pick_files)
        layout.addWidget(self.attach_btn, 0, Qt.AlignBottom)

        self.input_box = SmartTextEdit()
        self.input_box.setObjectName("inputBox")
        self.input_box.setFont(QFont("Georgia", 11))
        self.input_box.setPlaceholderText("Ask Digi Valet anything…")
        self.input_box.send_requested.connect(self._send_message)
        layout.addWidget(self.input_box)

        self.send_btn = QPushButton("↑")
        self.send_btn.setObjectName("sendBtn")
        self.send_btn.setFont(QFont("Georgia", 16, QFont.Bold))
        self.send_btn.setFixedSize(40, 40)
        self.send_btn.setToolTip("Send (Enter)")
        self.send_btn.clicked.connect(self._send_message)
        layout.addWidget(self.send_btn, 0, Qt.AlignBottom)

        self.stop_btn = QPushButton("■")
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setFont(QFont("Georgia", 12))
        self.stop_btn.setFixedSize(40, 40)
        self.stop_btn.setToolTip("Stop generation")
        self.stop_btn.clicked.connect(self._stop_generation)
        self.stop_btn.hide()
        layout.addWidget(self.stop_btn, 0, Qt.AlignBottom)

        # Hard cap: belt-and-suspenders safety net on top of the Fixed
        # size policies above, so this area can never be stretched by a
        # parent layout — chips strip + pill + outer margins, no more.
        input_row.setMaximumHeight(SmartTextEdit.MAX_HEIGHT + 12)
        outer_layout.addWidget(input_row)
        margins = outer_layout.contentsMargins()
        outer.setMaximumHeight(
            self._chips_bar.maximumHeight()
            + input_row.maximumHeight()
            + outer_layout.spacing()
            + margins.top() + margins.bottom()
        )
        return outer

    # ── File attachment logic ──────────────────────────────────────────────

    # ── Limits ──────────────────────────────────────────────────────────────
    _MAX_FILES = 100
    _MAX_TOTAL_BYTES = 10 * 1024 * 1024 * 1024  # 10 GB

    def _pick_files(self):
        """Open a file dialog — any type, up to 100 files / 10 GB total."""
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Attach files — any type, up to 100 files / 10 GB total",
            str(Path.home()),
            "All Files (*.*)"
        )
        skipped_dup = 0
        skipped_limit = 0
        skipped_size = 0

        current_total = sum(
            Path(p).stat().st_size for p in self._attached_files if Path(p).exists()
        )

        for path in paths:
            if not path:
                continue
            if path in self._attached_files:
                skipped_dup += 1
                continue
            if len(self._attached_files) >= self._MAX_FILES:
                skipped_limit += 1
                continue
            try:
                file_size = Path(path).stat().st_size
            except OSError:
                file_size = 0
            if current_total + file_size > self._MAX_TOTAL_BYTES:
                skipped_size += 1
                continue
            self._attached_files.append(path)
            self._add_chip(path)
            current_total += file_size

        self._refresh_chips_bar()
        self._update_attach_badge()

        # Friendly warnings
        msgs = []
        if skipped_limit:
            msgs.append(f"{skipped_limit} file(s) skipped — 100-file limit reached.")
        if skipped_size:
            msgs.append(f"{skipped_size} file(s) skipped — 10 GB total size limit reached.")
        if msgs:
            QMessageBox.warning(self, "Attachment Limit", "\n".join(msgs))

    def _add_chip(self, path: str):
        chip = FileChip(path)
        chip.removed.connect(self._remove_file)
        # Insert before the stretch
        idx = self._chips_layout.count() - 1
        self._chips_layout.insertWidget(idx, chip)
        self._file_chips[path] = chip

    def _remove_file(self, path: str):
        if path in self._attached_files:
            self._attached_files.remove(path)
        if path in self._file_chips:
            chip = self._file_chips.pop(path)
            self._chips_layout.removeWidget(chip)
            chip.deleteLater()
        self._refresh_chips_bar()
        self._update_attach_badge()

    def _refresh_chips_bar(self):
        self._chips_bar.setVisible(bool(self._attached_files))

    def _update_attach_badge(self):
        """Show file count on the attach button; revert to '+' when empty."""
        n = len(self._attached_files)
        if n == 0:
            self.attach_btn.setText("+")
            self.attach_btn.setToolTip("Attach files (any type, up to 100 files / 10 GB)")
        else:
            self.attach_btn.setText(f"+{n}")
            size_mb = sum(
                Path(p).stat().st_size for p in self._attached_files if Path(p).exists()
            ) / (1024 * 1024)
            size_str = f"{size_mb:.1f} MB" if size_mb < 1024 else f"{size_mb/1024:.2f} GB"
            self.attach_btn.setToolTip(
                f"{n} file(s) attached ({size_str}) — click to add more"
            )

    # ── Theme ──────────────────────────────────────────────────────────────

    def _apply_theme(self):
        global _DARK_MODE
        app = QApplication.instance()
        app.setStyleSheet(DARK_STYLESHEET if _DARK_MODE else LIGHT_STYLESHEET)
        if hasattr(self, 'theme_btn'):
            self.theme_btn.setText("☀ Light" if _DARK_MODE else "☾ Dark")
        if hasattr(self, 'title_bar'):
            self.title_bar.apply_theme(_DARK_MODE)
        self.setWindowIcon(make_app_icon(_DARK_MODE))
        if self._tray:
            self._tray.setIcon(make_app_icon(_DARK_MODE))

    def _toggle_theme(self):
        global _DARK_MODE
        _DARK_MODE = not _DARK_MODE
        self._prefs["dark_mode"] = _DARK_MODE
        self._apply_theme()
        self._save_prefs()
        # Re-render visible bubbles
        for i in range(self.chat_layout.count()):
            w = self.chat_layout.itemAt(i).widget()
            if isinstance(w, MessageBubble):
                w._render(w._raw_text)
        # Re-theme any attached file cards
        for chip in self._file_chips.values():
            chip._apply_style()

    # ── Privacy Mode ───────────────────────────────────────────────────────

    def _toggle_privacy(self, checked: bool):
        self.privacy_mode = checked
        self.privacy_btn.setText("🔒 Private ON" if checked else "🔒 Private")
        status = "Privacy mode ON — chat will not be saved." if checked else "Privacy mode OFF — chat is being saved."
        self._add_system_notice(status)

    def _add_system_notice(self, text: str):
        notice = QLabel(f"ℹ  {text}")
        notice.setFont(QFont("Georgia", 9))
        notice.setStyleSheet("color: #6a9ab8; padding: 4px 24px; background: transparent;")
        notice.setAlignment(Qt.AlignCenter)
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, notice)
        self._scroll_to_bottom()

    # ── Tone / Language ────────────────────────────────────────────────────

    def _on_tone_changed(self, text: str):
        self._prefs["tone"] = text.lower()
        self._save_prefs()
        self._refresh_system_prompt()

    def _on_language_changed(self, text: str):
        self._prefs["language"] = text
        self._save_prefs()
        self._refresh_system_prompt()

    def _refresh_system_prompt(self, include_programming_kb: bool = True):
        tone = self._prefs.get("tone", "balanced")
        language = self._prefs.get("language", "English")
        new_prompt = build_system_prompt(
            tone, language, kb=self.kb, include_programming_kb=include_programming_kb,
            org_tree=self.org_tree, excel_intel=_excel_intel
        )
        if self.messages and self.messages[0].get("role") == "system":
            self.messages[0]["content"] = new_prompt
        if self.current_chat_id and self.current_chat_id in self.all_chats:
            self.all_chats[self.current_chat_id]["messages"] = self.messages
            if not self.privacy_mode:
                self._save_all_chats()

    def _get_system_prompt(self, include_programming_kb: bool = True) -> str:
        tone = self._prefs.get("tone", "balanced")
        language = self._prefs.get("language", "English")
        return build_system_prompt(
            tone, language, kb=self.kb, include_programming_kb=include_programming_kb,
            org_tree=self.org_tree, excel_intel=_excel_intel
        )

    # ── Quick Commands ─────────────────────────────────────────────────────

    def _run_quick_command(self, cmd: str):
        if cmd == "/kb":
            self.input_box.setPlainText("/kb list")
            self._send_message()
            return
        if cmd == "/digivalet":
            self.input_box.setPlainText("/digivalet help")
            self._send_message()
            return
        prompt = QUICK_COMMANDS.get(cmd, "")
        if not prompt:
            return
        # Inject task context for /tasks command
        if cmd == "/tasks":
            task_summary = self.task_widget.get_task_summary()
            prompt = f"Here is my current task list:\n{task_summary}\n\n" + prompt
        self.input_box.setPlainText(prompt)
        self._send_message()

    # ── /digivalet org-chart command ─────────────────────────────────────

    def _handle_digivalet_command(self, text: str) -> str:
        """Return a markdown answer for a /digivalet command using the embedded OrgTree."""
        query = text[len("/digivalet"):].strip()
        ql = query.lower()

        if not self.org_tree.loaded:
            return (
                "⚠️ The org tree hasn't been loaded yet.\n\n"
                "Place **OrgTree.csv** (columns: `Employee Number, Display Name, "
                "Job Title, Department, Location, Reportees Count`) next to "
                "`digi_valet_chat.py` and restart, or use `/digivalet reload`."
            )

        # ── reload ─────────────────────────────────────────────────────────
        if ql == "reload":
            self._load_org_tree()
            if self.org_tree.loaded:
                return f"✓ OrgTree reloaded — {len(self.org_tree.employees)} employees."
            return "⚠️ Reload failed — OrgTree.csv not found."

        # ── export ─────────────────────────────────────────────────────────
        if ql == "export":
            import csv as _csv
            out = Path.home() / "digi_valet_org_export.csv"
            try:
                with open(out, 'w', newline='', encoding='utf-8') as f:
                    w = _csv.DictWriter(f, fieldnames=['emp_number','name','title','dept','location','reports_count'])
                    w.writeheader()
                    w.writerows(self.org_tree.employees.values())
                return f"✓ Exported {len(self.org_tree.employees)} employees to `{out}`"
            except Exception as e:
                return f"⚠️ Export failed: {e}"

        # ── help / empty ────────────────────────────────────────────────────
        if not query or ql in ("help", "?"):
            stats = self.org_tree.stats()
            dept_lines = "\n".join(f"  • {d}: {c}" for d, c in list(stats["departments"].items())[:10])
            return (
                f"**Digi Valet Org Lookup** — {stats['total']} employees loaded\n\n"
                "**Commands:**\n"
                "• `/digivalet <name>` — employee profile\n"
                "• `/digivalet <emp number>` — e.g. `/digivalet PB001`\n"
                "• `/digivalet <keyword>` — search across all fields\n"
                "• `/digivalet dept:<name>` — list by department\n"
                "• `/digivalet loc:<city>` — list by location\n"
                "• `/digivalet stats` — org overview\n"
                "• `/digivalet reload` — hot-reload OrgTree.csv\n"
                "• `/digivalet export` — export all employees to CSV\n\n"
                f"Top departments:\n{dept_lines}"
            )

        # ── stats ──────────────────────────────────────────────────────────
        if ql == "stats":
            stats = self.org_tree.stats()
            dept_lines = "\n".join(f"| {d} | {c} |" for d, c in stats["departments"].items())
            loc_lines = "\n".join(f"| {l} | {c} |" for l, c in stats["locations"].items())
            return (
                f"**Org Tree Overview** — {stats['total']} employees\n\n"
                f"**By Department**\n\n| Department | Count |\n|---|---|\n{dept_lines}\n\n"
                f"**By Location**\n\n| Location | Count |\n|---|---|\n{loc_lines}"
            )

        # ── dept: filter ────────────────────────────────────────────────────
        if ql.startswith("dept:"):
            dept_name = query[5:].strip()
            results = self.org_tree.list_by_department(dept_name, limit=50)
            if not results:
                return f"No employees found in department **'{dept_name}'**."
            return (
                f"**{len(results)} employee(s) in '{dept_name}':**\n\n"
                + self.org_tree.format_results_table(results)
            )

        # ── loc: filter ─────────────────────────────────────────────────────
        if ql.startswith("loc:"):
            loc_name = query[4:].strip()
            results = self.org_tree.list_by_location(loc_name, limit=50)
            if not results:
                return f"No employees found at location **'{loc_name}'**."
            return (
                f"**{len(results)} employee(s) at '{loc_name}':**\n\n"
                + self.org_tree.format_results_table(results)
            )

        # ── exact match (name or emp number) ────────────────────────────────
        emp = self.org_tree.get_employee(query)
        if emp:
            return self.org_tree.format_employee(emp)

        # ── full-text search ─────────────────────────────────────────────────
        results = self.org_tree.search(query)
        if not results:
            return f"No matches found in the org tree for **'{query}'**."
        return f"Found **{len(results)}** match(es) for **'{query}'**:\n\n" + self.org_tree.format_results_table(results)

    # ── State and History ──────────────────────────────────────────────────

    def _initialize_default_chat(self):
        self._populate_sidebar_list()
        if self.all_chats:
            sorted_keys = sorted(self.all_chats.keys(), reverse=True)
            self._switch_to_chat(sorted_keys[0])
        else:
            self._new_chat()

    def _populate_sidebar_list(self):
        self.history_list_widget.clear()
        for chat_id in sorted(self.all_chats.keys(), reverse=True):
            item = QListWidgetItem(self.all_chats[chat_id]["title"])
            item.setData(Qt.UserRole, chat_id)
            self.history_list_widget.addItem(item)
            if chat_id == self.current_chat_id:
                self.history_list_widget.setCurrentItem(item)

    def _switch_to_chat(self, chat_id):
        if self.current_worker:
            self.current_worker.cancel()
        self.current_chat_id = chat_id
        self.messages = self.all_chats[chat_id]["messages"]
        self._clear_chat_display()
        if len(self.messages) > 1:
            self._restore_chat_display()
        else:
            self._add_welcome()
        for i in range(self.history_list_widget.count()):
            item = self.history_list_widget.item(i)
            if item.data(Qt.UserRole) == chat_id:
                self.history_list_widget.setCurrentItem(item)
                break

    def _new_chat(self):
        if self.current_worker:
            self.current_worker.cancel()
        self.current_chat_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        self.messages = [{"role": "system", "content": self._get_system_prompt()}]
        self.all_chats[self.current_chat_id] = {
            "title": "New Conversation",
            "messages": self.messages
        }
        self._clear_chat_display()
        self._add_welcome()
        self._populate_sidebar_list()
        if not self.privacy_mode:
            self._save_all_chats()

    def _on_history_item_clicked(self, item):
        chat_id = item.data(Qt.UserRole)
        if chat_id != self.current_chat_id:
            self._switch_to_chat(chat_id)

    def _clear_chat_display(self):
        while self.chat_layout.count() > 1:
            item = self.chat_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    def _add_welcome(self):
        bubble = MessageBubble(
            "Good day. I'm Digi Valet, your personal AI assistant. "
            "How may I be of service today?\n\n"
            "**Tip:** Use the quick-command buttons above (e.g. `/help`, `/tasks`, `/plan`) for fast access to common features.",
            "assistant"
        )
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, bubble)

    def _restore_chat_display(self):
        for msg in self.messages[1:]:
            bubble = MessageBubble(msg.get("content", ""), msg.get("role", "user"))
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, bubble)
        self._scroll_to_bottom()

    # ── Model Loading ──────────────────────────────────────────────────────

    def _load_models(self):
        self.sidebar_status.setText("● Connecting…")
        thread = QThread(self)
        worker = OllamaModelsWorker(self.ollama_host)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.models_ready.connect(self._on_models_ready)
        worker.error.connect(self._on_models_error)
        worker.models_ready.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.start()
        self._model_thread = thread
        self._model_worker = worker

    def _on_models_ready(self, models):
        self.model_combo.clear()
        if models:
            for m in models:
                self.model_combo.addItem(m)
            self.sidebar_status.setText("● Ollama connected")
            self.sidebar_status.setStyleSheet("color: #6a9e6a; font-size: 10px;")
            self.model_indicator.setText(self.model_combo.currentText())
            self.model_combo.currentTextChanged.connect(
                lambda t: self.model_indicator.setText(t)
            )
        else:
            self.model_combo.addItem("No models found")
            self.sidebar_status.setText("● No models installed")
            self.sidebar_status.setStyleSheet("color: #c8a06e; font-size: 10px;")

    def _on_models_error(self, err):
        self.model_combo.clear()
        self.model_combo.addItem("Ollama not running")
        self.sidebar_status.setText("● Ollama offline")
        self.sidebar_status.setStyleSheet("color: #e07070; font-size: 10px;")

    # ── Core Generation Pipeline ───────────────────────────────────────────

    def _send_message(self):
        text = self.input_box.toPlainText().strip()
        if not text:
            return
        if self.current_thread and self.current_thread.isRunning():
            self._add_system_notice("⏳ Please wait for the current response to finish.")
            return

        # Resume auto-scroll on every new message
        self._auto_scroll = True

        # ── /excel command: answer locally ─────────────────────────────────
        if text.lower().startswith("/excel"):
            self.input_box.clear()
            user_bubble = MessageBubble(text, "user")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, user_bubble)
            answer = self._handle_excel_command(text)
            reply_bubble = MessageBubble(answer, "assistant")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, reply_bubble)
            self._scroll_to_bottom()
            self.messages.append({"role": "user", "content": text})
            self.messages.append({"role": "assistant", "content": answer})
            self.all_chats[self.current_chat_id]["messages"] = self.messages
            self._populate_sidebar_list()
            if not self.privacy_mode:
                self._save_all_chats()
            return

        # ── /kb knowledge-base command: answer locally, no LLM call ────
        if text.lower().startswith("/kb"):
            self.input_box.clear()
            user_bubble = MessageBubble(text, "user")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, user_bubble)
            answer = self._handle_kb_command(text)
            reply_bubble = MessageBubble(answer, "assistant")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, reply_bubble)
            self._scroll_to_bottom()
            self.messages.append({"role": "user",      "content": text})
            self.messages.append({"role": "assistant", "content": answer})
            self.all_chats[self.current_chat_id]["messages"] = self.messages
            self._populate_sidebar_list()
            if not self.privacy_mode:
                self._save_all_chats()
            return

        # ── /digivalet org-chart command: answer locally, no LLM call ──
        if text.lower().startswith("/digivalet"):
            self.input_box.clear()

            user_bubble = MessageBubble(text, "user")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, user_bubble)

            answer = self._handle_digivalet_command(text)
            reply_bubble = MessageBubble(answer, "assistant")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, reply_bubble)
            self._scroll_to_bottom()

            self.messages.append({"role": "user", "content": text})
            self.messages.append({"role": "assistant", "content": answer})
            self.all_chats[self.current_chat_id]["messages"] = self.messages
            self._populate_sidebar_list()
            if not self.privacy_mode:
                self._save_all_chats()
            return

        # ── Excel Intelligence: graph generation ────────────────────────────
        if _excel_intel and _excel_intel.has_data():
            graph_path = _excel_intel.try_generate_graph(text)
            if graph_path:
                self.input_box.clear()
                user_bubble = MessageBubble(text, "user")
                self.chat_layout.insertWidget(self.chat_layout.count() - 1, user_bubble)
                self._show_graph_in_chat(graph_path)
                self.messages.append({"role": "user", "content": text})
                self.messages.append({"role": "assistant", "content": f"📊 Graph generated: {graph_path}"})
                self.all_chats[self.current_chat_id]["messages"] = self.messages
                self._populate_sidebar_list()
                if not self.privacy_mode:
                    self._save_all_chats()
                return
        # ── End Excel Intelligence ───────────────────────────────────────────

        model = self.model_combo.currentText()
        if not model or model in ("Loading models…", "No models found", "Ollama not running"):
            self._add_error_bubble("Please select a valid Ollama model first.")
            return

        if len(self.messages) == 1 and self.all_chats.get(self.current_chat_id, {}).get("title") == "New Conversation":
            short_title = text[:24] + "..." if len(text) > 24 else text
            self.all_chats[self.current_chat_id]["title"] = short_title

        self.input_box.clear()

        # ── Build user message with file context injected ────────────────
        files_to_send = list(self._attached_files)   # snapshot before clearing

        # ── Auto-learn: if user says "learn this", store files permanently ──
        _learn_triggers = (
            "learn this", "remember this", "store this", "memorize this",
            "memorise this", "add to knowledge", "save to knowledge",
            "keep this", "add to kb",
        )
        if files_to_send and self.kb and any(t in text.lower() for t in _learn_triggers):
            learn_results = []
            for fpath in files_to_send:
                result = self.kb.learn_file(fpath)
                learn_results.append(result)
            # Immediately update system prompt with new KB data
            self._refresh_system_prompt()
            self._add_kb_notice("\n".join(learn_results))
        if files_to_send:
            file_sections = []
            for fpath in files_to_send:
                _, extracted = _read_file_content(fpath)
                file_sections.append(extracted)
            file_context = (
                "\n\n---\n**Attached files for context:**\n\n"
                + "\n\n---\n".join(file_sections)
                + "\n---"
            )
            full_user_text = text + file_context
        else:
            full_user_text = text

        # ── Protect context budget: if this turn carries real code (an
        # attachment or a pasted code block), drop the ~50KB generic
        # programming-syntax reference from the system prompt so the
        # model's limited context goes toward the user's actual code
        # instead of being silently truncated away.
        self._refresh_system_prompt(
            include_programming_kb=not _message_likely_contains_code(full_user_text)
        )

        user_bubble = MessageBubble(text, "user")   # show only the typed text in the bubble
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, user_bubble)

        # Clear file chips immediately after capturing them
        if files_to_send:
            for path in list(self._attached_files):
                self._remove_file(path)
            self._update_attach_badge()

        self.typing_indicator = TypingIndicator()
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, self.typing_indicator)
        self._scroll_to_bottom()

        self.messages.append({"role": "user", "content": full_user_text})
        self.all_chats[self.current_chat_id]["messages"] = self.messages
        self._populate_sidebar_list()
        if not self.privacy_mode:
            self._save_all_chats()

        self.current_bubble = None
        self.send_btn.hide()
        self.stop_btn.show()
        self.input_box.setEnabled(False)

        self.current_thread = QThread(self)
        self.current_worker = OllamaWorker(
            model, self._windowed_messages(), self.ollama_host
        )
        self.current_worker.moveToThread(self.current_thread)
        self.current_thread.started.connect(self.current_worker.run)
        self.current_worker.token_received.connect(self._on_token)
        self.current_worker.finished.connect(self._on_finished)
        self.current_worker.error.connect(self._on_error)
        self.current_thread.start()

    # ── History windowing ─────────────────────────────────────────────────

    _MAX_HISTORY_TURNS = 12  # user+assistant message pairs sent to the model

    def _windowed_messages(self) -> list:
        """Return system prompt + the most recent N turns.

        self.messages keeps the FULL conversation for display, export,
        and saved history. But sending the entire unbounded history to
        Ollama on every turn means old large pastes (or just a long
        conversation) can silently crowd out the current message within
        the model's context window — the same failure mode as the
        oversized knowledge-base injection. This caps what's actually
        sent, while leaving the saved/displayed history untouched.
        """
        if not self.messages:
            return self.messages
        system_msg = self.messages[0:1] if self.messages[0].get("role") == "system" else []
        rest = self.messages[1:] if system_msg else self.messages
        max_messages = self._MAX_HISTORY_TURNS * 2
        if len(rest) > max_messages:
            rest = rest[-max_messages:]
        return system_msg + rest

    def _on_token(self, token: str):
        if self.typing_indicator:
            try:
                self.typing_indicator.stop()
            except RuntimeError:
                pass
            self.chat_layout.removeWidget(self.typing_indicator)
            self.typing_indicator.deleteLater()
            self.typing_indicator = None

        if self.current_bubble is None:
            self.current_bubble = MessageBubble("", "assistant")
            self.chat_layout.insertWidget(self.chat_layout.count() - 1, self.current_bubble)

        self.current_bubble.append_text(token)
        self._scroll_to_bottom()

    def _on_finished(self):
        if self.typing_indicator:
            try:
                self.typing_indicator.stop()
            except RuntimeError:
                pass
            self.chat_layout.removeWidget(self.typing_indicator)
            self.typing_indicator.deleteLater()
            self.typing_indicator = None

        if self.current_bubble:
            full_text = self.current_bubble._raw_text
            self.messages.append({"role": "assistant", "content": full_text})
            self.all_chats[self.current_chat_id]["messages"] = self.messages
            if not self.privacy_mode:
                self._save_all_chats()

        self.send_btn.show()
        self.stop_btn.hide()
        self.input_box.setEnabled(True)
        self.input_box.setFocus()
        self.current_bubble = None
        self.current_worker = None
        if self.current_thread:
            self.current_thread.quit()
            self.current_thread.wait(2000)   # wait up to 2 s for thread to exit cleanly
            self.current_thread.deleteLater()
            self.current_thread = None

    def _on_error(self, err: str):
        if self.typing_indicator:
            try:
                self.typing_indicator.stop()
            except RuntimeError:
                pass
            self.chat_layout.removeWidget(self.typing_indicator)
            self.typing_indicator.deleteLater()
            self.typing_indicator = None
        self._add_error_bubble(err)
        self._on_finished()

    def _stop_generation(self):
        if self.current_worker:
            self.current_worker.cancel()

    def _add_error_bubble(self, text: str):
        bubble = MessageBubble(f"⚠ {text}", "assistant")
        bubble.setStyleSheet("""
            QFrame {
                background-color: #1a0e0e;
                border: 1px solid #5a2828;
                border-left: 3px solid #e07070;
                border-radius: 10px;
                margin: 3px 60px 3px 0px;
            }
        """)
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, bubble)
        self._scroll_to_bottom()

    def _clear_all_chats(self):
        self.all_chats = {}
        try:
            if self.chat_history_file.exists():
                self.chat_history_file.unlink()
        except Exception as e:
            print(f"Error purging file: {e}")
        self._new_chat()

    # ── Knowledge Base Commands ────────────────────────────────────────────

    def _handle_kb_command(self, text: str) -> str:
        """Handle /kb commands for the persistent knowledge base."""
        if not self.kb:
            return (
                "⚠️ Knowledge Base module not found.\n\n"
                "Make sure **knowledge_base.py** is in the same folder as "
                "digi_valet_chat.py, then restart Digi Valet."
            )
        parts = text.strip().split(None, 2)
        sub = parts[1].lower() if len(parts) > 1 else "list"

        if sub in ("list", "status", "ls", "show"):
            return self.kb.format_kb_status()

        if sub == "forget" and len(parts) > 2:
            result = self.kb.forget_file(parts[2].strip())
            self._refresh_system_prompt()
            return result

        if sub == "forget":
            return "Usage: `/kb forget <filename>`\nExample: `/kb forget OrgTree.csv`"

        if sub in ("learn", "add", "store"):
            if self._attached_files:
                results = []
                for fpath in list(self._attached_files):
                    results.append(self.kb.learn_file(fpath))
                self._refresh_system_prompt()
                return "\n\n".join(results)
            return (
                "No files attached.\n\n"
                "Attach a file first using the **+** button, then type `/kb learn`.\n"
                "Or attach a file and say **'learn this file'**."
            )

        if sub == "clear":
            files = self.kb.list_files()
            count = len(files)
            if count == 0:
                return "Knowledge base is already empty."
            for f in files:
                self.kb.forget_file(f["filename"])
            self._refresh_system_prompt()
            return f"🗑️ Cleared {count} file(s) from knowledge base."

        if sub == "help":
            return (
                "📚 **Knowledge Base Commands**\n\n"
                "• `/kb list` — show all stored files\n"
                "• `/kb learn` — store attached file(s) permanently *(attach first)*\n"
                "• `/kb forget <filename>` — remove a file\n"
                "• `/kb clear` — wipe the entire knowledge base\n"
                "• `/kb help` — show this help\n\n"
                "**Shortcut:** Attach any file and type **'learn this file'** — "
                "the model will remember it in all future conversations."
            )

        return self.kb.format_kb_status()

    def _add_kb_notice(self, text: str):
        """Show a green knowledge-base status notice in the chat."""
        notice = QLabel(f"📚  {text}")
        notice.setFont(QFont("Georgia", 9))
        notice.setWordWrap(True)
        notice.setStyleSheet(
            "color: #6a9e6a; padding: 6px 24px; background: transparent;"
        )
        notice.setAlignment(Qt.AlignLeft)
        self.chat_layout.insertWidget(self.chat_layout.count() - 1, notice)
        self._scroll_to_bottom()

    def _on_scroll_value_changed(self, value: int):
        """If the user drags the bar away from the bottom, pause auto-scroll."""
        sb = self.scroll_area.verticalScrollBar()
        at_bottom = (value >= sb.maximum() - 30)   # 30px tolerance
        if not at_bottom:
            self._auto_scroll = False
        else:
            self._auto_scroll = True

    def _scroll_to_bottom(self):
        """Only scroll if auto-scroll is enabled (user hasn't scrolled up)."""
        if self._auto_scroll:
            QTimer.singleShot(30, lambda: self.scroll_area.verticalScrollBar().setValue(
                self.scroll_area.verticalScrollBar().maximum()
            ))

    # ── Excel Graph Display ──────────────────────────────────────────────────

    def _show_graph_in_chat(self, graph_path: str):
        """Display a generated graph PNG inline in the chat area."""
        px = QPixmap(graph_path)
        if px.isNull():
            self._add_error_bubble(f"Graph saved but could not display: {graph_path}")
            return
        max_w = min(700, self.scroll_area.width() - 40)
        if px.width() > max_w:
            px = px.scaledToWidth(max_w, Qt.TransformationMode.SmoothTransformation)

        container = QFrame()
        container.setObjectName("assistantBubble")
        vl = QVBoxLayout(container)
        vl.setContentsMargins(10, 10, 10, 10)
        vl.setSpacing(6)

        hdr = QLabel("📊  Graph")
        hdr.setFont(QFont("Georgia", 9, QFont.Bold))
        hdr.setStyleSheet("color: #c8a96e;" if _DARK_MODE else "color: #8b6914;")
        vl.addWidget(hdr)

        img_label = QLabel()
        img_label.setPixmap(px)
        img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        vl.addWidget(img_label)

        path_label = QLabel(f"Saved: {graph_path}")
        path_label.setFont(QFont("Georgia", 7))
        path_label.setStyleSheet("color: #5a5048;" if _DARK_MODE else "color: #a09080;")
        vl.addWidget(path_label)

        self.chat_layout.insertWidget(self.chat_layout.count() - 1, container)
        QTimer.singleShot(60, lambda: self._scroll_to_bottom())

    def _handle_excel_command(self, text: str) -> str:
        """Handle /excel status command."""
        if _excel_intel is None:
            return ("⚠️ Excel Intelligence not available.\n\n"
                    "Install it: `pip install matplotlib openpyxl pandas`\n"
                    "Then place `excel_intelligence.py` next to this file.")
        return _excel_intel.get_db_stats()

    # ── Search / Rename ────────────────────────────────────────────────────

    def _filter_sidebar(self, query: str):
        query = query.strip().lower()
        for i in range(self.history_list_widget.count()):
            item = self.history_list_widget.item(i)
            title = item.text().lower()
            item.setHidden(bool(query) and query not in title)

    def _on_history_item_double_clicked(self, item):
        chat_id = item.data(Qt.UserRole)
        current_title = self.all_chats[chat_id]["title"]
        new_title, ok = QInputDialog.getText(
            self, "Rename Conversation", "New name:", text=current_title
        )
        if ok and new_title.strip():
            self.all_chats[chat_id]["title"] = new_title.strip()
            self._populate_sidebar_list()
            if not self.privacy_mode:
                self._save_all_chats()

    # ── Font Size ──────────────────────────────────────────────────────────

    def _change_font_size(self, delta: int):
        global _FONT_SIZE_DELTA
        _FONT_SIZE_DELTA = max(-4, min(6, _FONT_SIZE_DELTA + delta))
        self._font_size_delta = _FONT_SIZE_DELTA
        for i in range(self.chat_layout.count()):
            widget = self.chat_layout.itemAt(i).widget()
            if isinstance(widget, MessageBubble):
                widget.update_font_size(_FONT_SIZE_DELTA)
        base = 13 + _FONT_SIZE_DELTA
        self.input_box.setFont(QFont("Georgia", max(9, base)))

    # ── Export ─────────────────────────────────────────────────────────────

    def _export_chat(self):
        if not self.messages or len(self.messages) <= 1:
            QMessageBox.information(self, "Export", "No messages to export yet.")
            return

        title = self.all_chats.get(self.current_chat_id, {}).get("title", "conversation")
        safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')
        if not safe_title:
            safe_title = "conversation"

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Conversation",
            str(Path.home() / f"{safe_title}.md"),
            "Markdown (*.md);;Plain Text (*.txt)"
        )
        if not path:
            return

        lines = [f"# {title}\n", f"*Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}*\n\n---\n"]
        for msg in self.messages:
            role = msg.get("role", "")
            content = msg.get("content", "")
            if role == "system":
                continue
            speaker = "**Digi Valet**" if role == "assistant" else "**You**"
            lines.append(f"{speaker}\n\n{content}\n\n---\n")

        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            QMessageBox.information(self, "Export Complete", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    # ── Persistence ────────────────────────────────────────────────────────

    def _save_all_chats(self):
        try:
            with open(self.chat_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.all_chats, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving chat history: {e}")

    def _load_all_chats(self):
        try:
            if self.chat_history_file.exists():
                with open(self.chat_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading chat history: {e}")
        return {}

    def closeEvent(self, event):
        """Minimise to the system tray instead of quitting (unless Quit was chosen)."""
        if self._force_quit or not self._tray or not self._tray.isVisible():
            self._do_quit()
            event.accept()
            return

        event.ignore()
        self.hide()
        self._tray.showMessage(
            "Digi Valet",
            "Still running in the system tray. Double-click the tray icon to bring it back.",
            QSystemTrayIcon.MessageIcon.Information,
            2500,
        )

    def _do_quit(self):
        if self.current_worker:
            self.current_worker.cancel()
        if not self.privacy_mode:
            self._save_all_chats()
        self._save_prefs()
        if self._tray:
            self._tray.hide()

    # ── System Tray ───────────────────────────────────────────────────────

    def _setup_tray(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return

        self._tray = QSystemTrayIcon(self)
        self._tray.setIcon(make_app_icon(_DARK_MODE))
        self._tray.setToolTip("Digi Valet")
        self._tray.activated.connect(self._on_tray_activated)

        menu = QMenu()
        act_show = QAction("◈  Show Digi Valet", self)
        act_show.triggered.connect(self._show_from_tray)
        menu.addAction(act_show)

        act_new = QAction("＋  New Conversation", self)
        act_new.triggered.connect(self._new_chat)
        menu.addAction(act_new)

        act_theme = QAction("☀／☾  Toggle Theme", self)
        act_theme.triggered.connect(self._toggle_theme)
        menu.addAction(act_theme)

        menu.addSeparator()

        act_quit = QAction("✕  Quit", self)
        act_quit.triggered.connect(self._quit_app)
        menu.addAction(act_quit)

        self._tray.setContextMenu(menu)
        self._tray.show()

    def _on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.ActivationReason.DoubleClick:
            self._show_from_tray()

    def _show_from_tray(self):
        self.show()
        self.raise_()
        self.activateWindow()

    def _quit_app(self):
        self._force_quit = True
        self.close()
        QApplication.instance().quit()

    # ── Keyboard Shortcuts ────────────────────────────────────────────────

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("F11"), self, self._toggle_fullscreen)
        QShortcut(QKeySequence("Escape"), self, self._exit_fullscreen)
        QShortcut(QKeySequence("Ctrl+W"), self, self.close)
        QShortcut(QKeySequence("Ctrl+Q"), self, self._quit_app)
        QShortcut(QKeySequence("Ctrl+N"), self, self._new_chat)

    def _toggle_fullscreen(self):
        if self.isFullScreen():
            self.showNormal()
        else:
            self.showFullScreen()

    def _exit_fullscreen(self):
        if self.isFullScreen():
            self.showNormal()


# ─── Entry Point ───────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Digi Valet")
    app.setQuitOnLastWindowClosed(False)
    prefs_file = Path.home() / ".digi_valet_prefs.json"
    dark = True
    try:
        if prefs_file.exists():
            dark = json.loads(prefs_file.read_text()).get("dark_mode", True)
    except Exception:
        pass
    app.setStyleSheet(DARK_STYLESHEET if dark else LIGHT_STYLESHEET)

    window = DigiValetWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()